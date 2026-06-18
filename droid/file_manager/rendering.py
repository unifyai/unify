"""
Document Rendering utilities.

Provides functions to render documents as PIL Images:
- render_excel_sheet: Render Excel sheet ranges with styling, charts, and images
- render_pdf: Render PDF pages as images

Copied from examplecorp_data_agent/src/rendering.py
"""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import TYPE_CHECKING
from colorsys import rgb_to_hls, hls_to_rgb
from datetime import datetime
from io import BytesIO

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from PIL import Image, ImageDraw, ImageFont

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook import Workbook
    from openpyxl.cell import Cell
    import pymupdf

# ============================================================================
# Constants
# ============================================================================

DEFAULT_FONT_SIZE = 11
HEADER_BG_COLOR = (242, 242, 242)
GRID_COLOR = (200, 200, 200)
HEADER_BORDER_COLOR = (150, 150, 150)
DEFAULT_CELL_BG = (255, 255, 255)
DEFAULT_TEXT_COLOR = (0, 0, 0)

DEFAULT_COL_WIDTH = 64
DEFAULT_ROW_HEIGHT = 20
ROW_HEADER_WIDTH = 40
COL_HEADER_HEIGHT = 20
CELL_PADDING = 3

EMU_PER_PIXEL = 914400 / 96

BORDER_WIDTHS = {
    "thin": 1,
    "medium": 2,
    "thick": 3,
    "double": 3,
    "hair": 1,
    "dotted": 1,
    "dashed": 1,
    "mediumDashed": 2,
}

EXCEL_COLORS = [
    "#4472C4",
    "#ED7D31",
    "#A5A5A5",
    "#FFC000",
    "#5B9BD5",
    "#70AD47",
    "#264478",
    "#9E480E",
    "#636363",
    "#997300",
    "#255E91",
    "#43682B",
]

INDEXED_COLORS = {
    0: "000000",
    1: "FFFFFF",
    2: "FF0000",
    3: "00FF00",
    4: "0000FF",
    5: "FFFF00",
    6: "FF00FF",
    7: "00FFFF",
    8: "000000",
    9: "FFFFFF",
    10: "FF0000",
    11: "00FF00",
    12: "0000FF",
    13: "FFFF00",
    14: "FF00FF",
    15: "00FFFF",
    16: "800000",
    17: "008000",
    18: "000080",
    19: "808000",
    20: "800080",
    21: "008080",
    22: "C0C0C0",
    23: "808080",
    24: "9999FF",
    25: "993366",
    26: "FFFFCC",
    27: "CCFFFF",
    28: "660066",
    29: "FF8080",
    30: "0066CC",
    31: "CCCCFF",
    32: "000080",
    33: "FF00FF",
    34: "FFFF00",
    35: "00FFFF",
    36: "800080",
    37: "800000",
    38: "008080",
    39: "0000FF",
    40: "00CCFF",
    41: "CCFFFF",
    42: "CCFFCC",
    43: "FFFF99",
    44: "99CCFF",
    45: "FF99CC",
    46: "CC99FF",
    47: "FFCC99",
    48: "3366FF",
    49: "33CCCC",
    50: "99CC00",
    51: "FFCC00",
    52: "FF9900",
    53: "FF6600",
    54: "666699",
    55: "969696",
    56: "003366",
    57: "339966",
    58: "003300",
    59: "333300",
    60: "993300",
    61: "993366",
    62: "333399",
    63: "333333",
    64: "000000",
}

_font_cache: dict[tuple, ImageFont.FreeTypeFont] = {}

# ============================================================================
# Utility Functions
# ============================================================================


def parse_range(range_str: str) -> tuple[str, int, str, int]:
    """Parse 'A1:C10' into (start_col, start_row, end_col, end_row)."""
    match = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", range_str)
    if not match:
        raise ValueError(f"Invalid range: {range_str}")
    return (
        match.group(1).upper(),
        int(match.group(2)),
        match.group(3).upper(),
        int(match.group(4)),
    )


def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    """Convert hex color to RGB tuple."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 8:  # ARGB format
        hex_color = hex_color[2:]
    return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))


def apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    """Apply Excel tint transformation to RGB color."""
    r, g, b = rgb[0] / 255.0, rgb[1] / 255.0, rgb[2] / 255.0
    h, l, s = rgb_to_hls(r, g, b)
    l = l * (1 + tint) if tint < 0 else l + (1 - l) * tint
    l = max(0, min(1, l))
    r, g, b = hls_to_rgb(h, l, s)
    return (int(r * 255), int(g * 255), int(b * 255))


def get_column_width_pixels(ws: Worksheet, col_letter: str) -> int:
    col_dim = ws.column_dimensions.get(col_letter)
    return (
        int(col_dim.width * 7 + 5) if col_dim and col_dim.width else DEFAULT_COL_WIDTH
    )


def get_row_height_pixels(ws: Worksheet, row_num: int) -> int:
    row_dim = ws.row_dimensions.get(row_num)
    return (
        int(row_dim.height * 1.33) if row_dim and row_dim.height else DEFAULT_ROW_HEIGHT
    )


def get_default_font(
    size: int = DEFAULT_FONT_SIZE,
    bold: bool = False,
    italic: bool = False,
    scale: float = 1.0,
) -> ImageFont.FreeTypeFont:
    """Get a cached font with optional scaling."""
    scaled_size = int(size * scale)
    key = (scaled_size, bold, italic)
    if key not in _font_cache:
        # Include Unicode-capable fonts for symbols like checkmarks
        if bold:
            font_names = [
                "arialbd.ttf",
                "arial.ttf",
                "seguisym.ttf",
                "segoeuib.ttf",
                "segoeui.ttf",
            ]
        else:
            font_names = [
                "arial.ttf",
                "arialbd.ttf",
                "seguisym.ttf",
                "segoeui.ttf",
                "segoeuib.ttf",
            ]

        for font_name in font_names:
            try:
                _font_cache[key] = ImageFont.truetype(font_name, scaled_size)
                break
            except OSError:
                pass
        if key not in _font_cache:
            _font_cache[key] = ImageFont.load_default()
    return _font_cache[key]


# Cache for symbol font (used for checkmarks and special characters)
_symbol_font_cache: dict[int, ImageFont.FreeTypeFont] = {}


def get_symbol_font(
    size: int = DEFAULT_FONT_SIZE,
    scale: float = 1.0,
) -> ImageFont.FreeTypeFont:
    """Get a Unicode-capable font for symbols like checkmarks."""
    scaled_size = int(size * scale)
    if scaled_size not in _symbol_font_cache:
        # Fonts known to have good Unicode symbol support
        symbol_fonts = ["seguisym.ttf", "segoeui.ttf", "DejaVuSans.ttf", "arial.ttf"]
        for font_name in symbol_fonts:
            try:
                _symbol_font_cache[scaled_size] = ImageFont.truetype(
                    font_name,
                    scaled_size,
                )
                break
            except OSError:
                pass
        if scaled_size not in _symbol_font_cache:
            _symbol_font_cache[scaled_size] = ImageFont.load_default()
    return _symbol_font_cache[scaled_size]


def has_special_symbols(text: str) -> bool:
    """Check if text contains special Unicode symbols that may need a symbol font."""
    # Common symbols that Arial doesn't render well
    special_chars = set("✓✔✗✘☑☐☒★☆●○◆◇▲△▼▽►◄")
    return any(c in special_chars for c in text)


def truncate_text_to_width(
    draw: ImageDraw.Draw,
    text: str,
    font: ImageFont.FreeTypeFont,
    max_width: int,
    ellipsis: str = "...",
) -> str:
    """Truncate text to fit within max_width, adding ellipsis if needed."""
    if max_width <= 0:
        return ""

    # Check if text already fits
    bbox = draw.textbbox((0, 0), text, font=font)
    if bbox[2] - bbox[0] <= max_width:
        return text

    # Calculate ellipsis width
    ellipsis_bbox = draw.textbbox((0, 0), ellipsis, font=font)
    ellipsis_width = ellipsis_bbox[2] - ellipsis_bbox[0]

    # If even ellipsis doesn't fit, return empty or partial ellipsis
    if ellipsis_width >= max_width:
        return ""

    # Binary search for the maximum text length that fits
    available_width = max_width - ellipsis_width
    left, right = 0, len(text)

    while left < right:
        mid = (left + right + 1) // 2
        truncated = text[:mid]
        bbox = draw.textbbox((0, 0), truncated, font=font)
        if bbox[2] - bbox[0] <= available_width:
            left = mid
        else:
            right = mid - 1

    if left == 0:
        return ellipsis if ellipsis_width <= max_width else ""

    return text[:left] + ellipsis


# ============================================================================
# Theme and Color Handling
# ============================================================================


def is_valid_hex_color(val: str) -> bool:
    """Check if a string is a valid 6-character hex color."""
    if len(val) != 6:
        return False
    try:
        int(val, 16)
        return True
    except ValueError:
        return False


def extract_theme_colors(wb: "Workbook") -> list[str]:
    """Extract theme colors from workbook."""
    colors = [
        "FFFFFF",
        "000000",
        "4472C4",
        "ED7D31",
        "A5A5A5",
        "FFC000",
        "5B9BD5",
        "70AD47",
        "7030A0",
        "FF0000",
        "00B050",
        "0070C0",
    ]
    try:
        theme_xml = wb.loaded_theme
        if not theme_xml:
            return colors
        root = ET.fromstring(theme_xml) if isinstance(theme_xml, bytes) else theme_xml
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}

        for scheme in root.findall(".//a:clrScheme", ns):
            color_map = [
                ("lt1", 0),
                ("dk1", 1),
                ("lt2", 2),
                ("dk2", 3),
                ("accent1", 4),
                ("accent2", 5),
                ("accent3", 6),
                ("accent4", 7),
                ("accent5", 8),
                ("accent6", 9),
                ("hlink", 10),
                ("folHlink", 11),
            ]
            for name, idx in color_map:
                elem = scheme.find(f"a:{name}", ns)
                if elem is not None:
                    for color_elem in elem:
                        # Check for lastClr attribute first (actual color value for system colors)
                        if "lastClr" in color_elem.attrib:
                            last_clr = color_elem.attrib["lastClr"]
                            if is_valid_hex_color(last_clr):
                                colors[idx] = last_clr
                                break
                        if "val" in color_elem.attrib:
                            val = color_elem.attrib["val"]
                            # Check for system color names first, before hex validation
                            if val == "window" or val == "windowBackground":
                                colors[idx] = "FFFFFF"
                                break
                            elif val == "windowText":
                                colors[idx] = "000000"
                                break
                            elif is_valid_hex_color(val):
                                colors[idx] = val
                                break
    except Exception:
        pass
    return colors


def get_color_from_cell_color(
    color_obj,
    theme_colors: list[str],
    is_background: bool = False,
) -> tuple[int, int, int] | None:
    """Resolve openpyxl color object to RGB tuple.

    Args:
        color_obj: openpyxl color object
        theme_colors: List of theme colors extracted from workbook
        is_background: If True, treat "automatic" colors (indexed 64/65) as no-fill
    """
    if not color_obj:
        return None
    try:
        if hasattr(color_obj, "rgb") and color_obj.rgb and color_obj.rgb != "00000000":
            if isinstance(color_obj.rgb, str) and len(color_obj.rgb) >= 6:
                return hex_to_rgb(color_obj.rgb)
        if hasattr(color_obj, "indexed") and color_obj.indexed is not None:
            # Indexed colors 64 and 65 are "System Foreground" and "System Background"
            # These are "automatic" colors - when used as cell backgrounds, they mean "no fill"
            if color_obj.indexed in (64, 65) and is_background:
                return None
            if color_obj.indexed in INDEXED_COLORS:
                return hex_to_rgb(INDEXED_COLORS[color_obj.indexed])
        if hasattr(color_obj, "theme") and color_obj.theme is not None:
            idx = color_obj.theme
            if 0 <= idx < len(theme_colors):
                rgb = hex_to_rgb(theme_colors[idx])
                if hasattr(color_obj, "tint") and color_obj.tint:
                    rgb = apply_tint(rgb, color_obj.tint)
                return rgb
    except Exception:
        pass
    return None


def get_cell_background(cell: "Cell", theme_colors: list[str]) -> tuple[int, int, int]:
    """Get cell background color."""
    if not hasattr(cell, "fill") or not cell.fill:
        return DEFAULT_CELL_BG

    fill = cell.fill

    # Only skip if explicitly no fill
    if not fill.fill_type or fill.fill_type == "none":
        return DEFAULT_CELL_BG

    # Try foreground color (used by solid fills and patterns)
    # Pass is_background=True to handle automatic/system colors properly
    color = get_color_from_cell_color(fill.fgColor, theme_colors, is_background=True)
    if color:
        return color

    # Try background color as fallback (some patterns use this)
    color = get_color_from_cell_color(fill.bgColor, theme_colors, is_background=True)
    if color:
        return color

    return DEFAULT_CELL_BG


def get_text_color(cell: "Cell", theme_colors: list[str]) -> tuple[int, int, int]:
    """Get cell text color."""
    if hasattr(cell, "font") and cell.font and cell.font.color:
        color = get_color_from_cell_color(cell.font.color, theme_colors)
        if color:
            return color
    return DEFAULT_TEXT_COLOR


# ============================================================================
# Number Formatting
# ============================================================================

FORMAT_COLOR_MAP = {
    "red": (255, 0, 0),
    "blue": (0, 0, 255),
    "green": (0, 128, 0),
    "yellow": (255, 255, 0),
    "cyan": (0, 255, 255),
    "magenta": (255, 0, 255),
}


def parse_format_sections(fmt: str) -> list[str]:
    """Split Excel format string into sections (positive; negative; zero; text)."""
    sections, current, depth = [], "", 0
    for c in fmt:
        if c == "[":
            depth += 1
        elif c == "]":
            depth -= 1
        if c == ";" and depth == 0:
            sections.append(current)
            current = ""
        else:
            current += c
    sections.append(current)
    return sections


def extract_format_color(section: str) -> tuple[int, int, int] | None:
    """Extract color like [Red] from format section."""
    match = re.search(
        r"\[(" + "|".join(FORMAT_COLOR_MAP.keys()) + r")\]",
        section,
        re.IGNORECASE,
    )
    return FORMAT_COLOR_MAP.get(match.group(1).lower()) if match else None


def format_value_with_color(cell: "Cell") -> tuple[str, tuple[int, int, int] | None]:
    """Format cell value and return (formatted_string, color_override)."""
    value = cell.value
    if value is None:
        return "", None

    fmt = cell.number_format or "General"

    # Date formatting
    if isinstance(value, datetime):
        return _format_date(value, fmt), None

    # Number formatting
    if isinstance(value, (int, float)):
        sections = parse_format_sections(fmt)
        # Determine section: positive=0, negative=1, zero=2
        idx = 1 if value < 0 else (2 if value == 0 and len(sections) > 2 else 0)
        section = sections[min(idx, len(sections) - 1)]

        color = extract_format_color(section)
        text = _format_number(value, section)
        return text, color

    return str(value), None


def _format_date(value: datetime, fmt: str) -> str:
    """Format datetime value."""
    fmt_lower = fmt.lower()
    if "d" in fmt_lower or "m" in fmt_lower or "y" in fmt_lower:
        py_fmt = fmt.replace("yyyy", "%Y").replace("yy", "%y").replace("mmmm", "%B")
        py_fmt = py_fmt.replace("mmm", "%b").replace("dd", "%d").replace("d", "%d")
        # Handle mm (month) vs mm (minute) - month comes before day
        if "mm" in py_fmt and "d" in fmt_lower:
            py_fmt = py_fmt.replace("mm", "%m", 1)
        try:
            return value.strftime(py_fmt)
        except:
            pass
    return value.strftime("%d/%m/%Y")


def _format_number(value: float, section: str) -> str:
    """Format number according to Excel format section."""
    abs_val = abs(value)

    # Handle "General" format - use Python's intelligent formatting
    if section.lower() == "general":
        # Use repr-like formatting that preserves significant digits
        if abs_val == 0:
            return "0"
        elif abs_val >= 1e10 or abs_val < 1e-4:
            # Scientific notation for very large/small numbers
            return f"{value:g}"
        else:
            # Regular formatting - remove trailing zeros
            formatted = f"{value:.10f}".rstrip("0").rstrip(".")
            return formatted

    # Percentage
    if "%" in section:
        decimals = section.count("0") - 1 if "." in section else 0
        return f"{abs_val * 100:,.{max(0, decimals)}f}%"

    # Currency symbols
    currency = ""
    for sym in ["£", "$", "€", "¥"]:
        if sym in section:
            currency = sym
            break

    # Parentheses for negatives
    use_parens = "(" in section and value < 0

    # Decimal places
    decimals = 0
    if "." in section:
        after_dot = section.split(".")[-1]
        decimals = len(re.findall(r"[0#]", after_dot.split(")")[0].split(";")[0]))

    # Format with thousands separator if comma in format
    if "," in section:
        formatted = f"{abs_val:,.{decimals}f}"
    else:
        formatted = f"{abs_val:.{decimals}f}"

    # Apply currency and parentheses
    if use_parens:
        return f"{currency}({formatted})"
    elif value < 0 and not use_parens:
        return f"-{currency}{formatted}" if currency else f"-{formatted}"
    return f"{currency}{formatted}"


# ============================================================================
# Font Handling
# ============================================================================


def get_font_for_cell(cell: "Cell", scale: float = 1.0) -> ImageFont.FreeTypeFont:
    """Get appropriate font for cell with optional scaling."""
    size = DEFAULT_FONT_SIZE
    bold = italic = False
    if hasattr(cell, "font") and cell.font:
        if cell.font.size:
            size = int(cell.font.size)
        bold = bool(cell.font.bold)
        italic = bool(cell.font.italic)
    return get_default_font(size, bold, italic, scale)


# ============================================================================
# Border Drawing
# ============================================================================


def draw_cell_borders(
    draw: ImageDraw.Draw,
    cell: "Cell",
    x: int,
    y: int,
    x2: int,
    y2: int,
    theme_colors: list[str],
    scale: float,
):
    """Draw cell borders."""
    if not hasattr(cell, "border") or not cell.border:
        return

    border = cell.border
    for side, coords in [
        ("top", [(x, y, x2, y)]),
        ("bottom", [(x, y2, x2, y2)]),
        ("left", [(x, y, x, y2)]),
        ("right", [(x2, y, x2, y2)]),
    ]:
        side_obj = getattr(border, side, None)
        if side_obj and side_obj.style:
            width = int(BORDER_WIDTHS.get(side_obj.style, 1) * scale)
            color = (
                get_color_from_cell_color(side_obj.color, theme_colors)
                if side_obj.color
                else (0, 0, 0)
            )
            draw.line(coords[0], fill=color, width=max(1, width))


# ============================================================================
# Chart Rendering
# ============================================================================


def _get_chart_title(chart) -> str:
    """Extract chart title."""
    try:
        if hasattr(chart, "title") and chart.title and hasattr(chart.title, "tx"):
            if hasattr(chart.title.tx, "rich"):
                for p in chart.title.tx.rich.p:
                    for r in p.r:
                        if hasattr(r, "t"):
                            return r.t
    except:
        pass
    return ""


def _extract_chart_data(chart, ws) -> tuple[list, list[list], list[str]]:
    """Extract (categories, series_data, series_names) from chart."""
    categories, series_data, series_names = [], [], []

    try:
        for series in getattr(chart, "series", []):
            values, name = [], ""

            # Series name
            if hasattr(series, "tx") and series.tx:
                if hasattr(series.tx, "strRef") and series.tx.strRef:
                    try:
                        ref = series.tx.strRef.f.split("!")[-1].replace("$", "")
                        name = str(ws[ref].value or "")
                    except:
                        pass
            series_names.append(name)

            # Data values
            if hasattr(series, "val") and series.val and hasattr(series.val, "numRef"):
                try:
                    ref = series.val.numRef.f.split("!")[-1].replace("$", "")
                    for cell in ws[ref]:
                        cells = cell if hasattr(cell, "__iter__") else [cell]
                        for c in cells:
                            values.append(
                                (
                                    float(c.value)
                                    if isinstance(c.value, (int, float))
                                    else 0
                                ),
                            )
                except:
                    pass
            series_data.append(values)

            # Categories from first series
            if not categories:
                for cat_attr in ["cat", "xVal"]:
                    cat_obj = getattr(series, cat_attr, None)
                    if cat_obj:
                        for ref_type in ["strRef", "numRef"]:
                            ref_obj = getattr(cat_obj, ref_type, None)
                            if ref_obj:
                                try:
                                    ref = ref_obj.f.split("!")[-1].replace("$", "")
                                    for cell in ws[ref]:
                                        cells = (
                                            cell
                                            if hasattr(cell, "__iter__")
                                            else [cell]
                                        )
                                        categories.extend(
                                            str(c.value or "") for c in cells
                                        )
                                except:
                                    pass
                                break
    except:
        pass

    return categories, series_data, series_names


def _render_chart_to_image(chart, ws, width: int, height: int) -> Image.Image | None:
    """Render chart to PIL Image using matplotlib."""
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np

        chart_type = type(chart).__name__
        title = _get_chart_title(chart)
        categories, series_data, series_names = _extract_chart_data(chart, ws)

        fig, ax = plt.subplots(figsize=(width / 100, height / 100), dpi=100)
        colors = EXCEL_COLORS
        rendered = False

        def setup_axis(ax, cats, names):
            if cats:
                ax.set_xticks(range(len(cats)))
                ax.set_xticklabels(cats, rotation=45, ha="right", fontsize=8)
            if len(names) > 1 or any(names):
                ax.legend(fontsize=7, loc="best")

        # Bar charts
        if chart_type in ("BarChart", "BarChart3D") and categories and series_data:
            x = np.arange(len(categories))
            w = 0.8 / max(len(series_data), 1)
            for i, (data, name) in enumerate(zip(series_data, series_names)):
                ax.bar(
                    x + (i - len(series_data) / 2 + 0.5) * w,
                    data[: len(categories)],
                    w,
                    label=name or f"Series {i+1}",
                    color=colors[i % len(colors)],
                )
            setup_axis(ax, categories, series_names)
            rendered = True

        # Line charts
        elif chart_type in ("LineChart", "LineChart3D") and categories and series_data:
            x = np.arange(len(categories))
            for i, (data, name) in enumerate(zip(series_data, series_names)):
                ax.plot(
                    x[: len(data)],
                    data,
                    "o-",
                    markersize=4,
                    linewidth=2,
                    label=name or f"Series {i+1}",
                    color=colors[i % len(colors)],
                )
            setup_axis(ax, categories, series_names)
            ax.grid(True, alpha=0.3)
            rendered = True

        # Area charts
        elif chart_type in ("AreaChart", "AreaChart3D") and categories and series_data:
            x = np.arange(len(categories))
            for i, (data, name) in enumerate(zip(series_data, series_names)):
                ax.fill_between(
                    x[: len(data)],
                    data,
                    alpha=0.5,
                    color=colors[i % len(colors)],
                    label=name or f"Series {i+1}",
                )
                ax.plot(
                    x[: len(data)],
                    data,
                    color=colors[i % len(colors)],
                    linewidth=1,
                )
            setup_axis(ax, categories, series_names)
            rendered = True

        # Pie/Doughnut charts
        elif chart_type in (
            "PieChart",
            "PieChart3D",
            "DoughnutChart",
            "ProjectedPieChart",
        ):
            if series_data and series_data[0]:
                data = [abs(v) for v in series_data[0]]
                labels = (
                    [str(c)[:15] for c in categories]
                    if categories
                    else [f"S{i+1}" for i in range(len(data))]
                )
                wedge_props = dict(width=0.5) if chart_type == "DoughnutChart" else {}
                ax.pie(
                    data,
                    labels=labels,
                    autopct="%1.0f%%",
                    colors=colors[: len(data)],
                    textprops={"fontsize": 7},
                    wedgeprops=wedge_props,
                )
                rendered = True

        # Scatter charts
        elif chart_type == "ScatterChart" and series_data:
            for i, (data, name) in enumerate(zip(series_data, series_names)):
                x_vals = list(range(len(data)))
                ax.scatter(
                    x_vals,
                    data,
                    s=50,
                    alpha=0.7,
                    color=colors[i % len(colors)],
                    label=name or f"Series {i+1}",
                )
            if any(series_names):
                ax.legend(fontsize=7)
            ax.grid(True, alpha=0.3)
            rendered = True

        # Radar charts
        elif chart_type == "RadarChart" and categories and series_data:
            angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
            angles += angles[:1]
            ax = fig.add_subplot(111, projection="polar")
            for i, (data, name) in enumerate(zip(series_data, series_names)):
                values = list(data[: len(categories)]) + [data[0]]
                ax.plot(
                    angles,
                    values,
                    "o-",
                    linewidth=2,
                    color=colors[i % len(colors)],
                    label=name or f"Series {i+1}",
                )
                ax.fill(angles, values, alpha=0.25, color=colors[i % len(colors)])
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(categories, fontsize=7)
            rendered = True

        # Fallback: bar chart
        if not rendered and series_data and series_data[0]:
            ax.bar(range(len(series_data[0])), series_data[0], color=colors[0])
            rendered = True

        if title:
            ax.set_title(title, fontsize=10, fontweight="bold")
        ax.tick_params(axis="both", labelsize=8)
        plt.tight_layout()

        buf = BytesIO()
        fig.savefig(buf, format="png", facecolor="white", edgecolor="none")
        buf.seek(0)
        plt.close(fig)
        return Image.open(buf).convert("RGB")
    except:
        return None


def _get_anchor_extent(anchor, ws=None) -> tuple[int, int, int, int, int, int]:
    """Get (col, row, width_px, height_px, extra_cols, extra_rows) from anchor.

    Handles both OneCellAnchor (has ext) and TwoCellAnchor (has to).
    """
    if not hasattr(anchor, "_from"):
        return 0, 0, 0, 0, 0, 0

    col = anchor._from.col + 1  # 0-indexed to 1-indexed
    row = anchor._from.row + 1

    # OneCellAnchor: size from ext
    if hasattr(anchor, "ext") and anchor.ext:
        w = int((anchor.ext.cx or 0) / EMU_PER_PIXEL)
        h = int((anchor.ext.cy or 0) / EMU_PER_PIXEL)

    # TwoCellAnchor: size from _from to to positions
    elif hasattr(anchor, "to") and anchor.to and ws:
        from_col, from_row = anchor._from.col, anchor._from.row
        to_col, to_row = anchor.to.col, anchor.to.row
        from_col_off = getattr(anchor._from, "colOff", 0) or 0
        from_row_off = getattr(anchor._from, "rowOff", 0) or 0
        to_col_off = getattr(anchor.to, "colOff", 0) or 0
        to_row_off = getattr(anchor.to, "rowOff", 0) or 0

        # Calculate width: sum of cell widths between columns + offset difference
        w = 0
        for c in range(from_col, to_col):
            w += get_column_width_pixels(ws, get_column_letter(c + 1))
        # Add the offset within the end cell, subtract offset in start cell
        w += int(to_col_off / EMU_PER_PIXEL) - int(from_col_off / EMU_PER_PIXEL)

        # Calculate height: sum of row heights between rows + offset difference
        h = 0
        for r in range(from_row, to_row):
            h += get_row_height_pixels(ws, r + 1)
        h += int(to_row_off / EMU_PER_PIXEL) - int(from_row_off / EMU_PER_PIXEL)

        # Ensure minimum size
        w = max(w, 10)
        h = max(h, 10)

    # TwoCellAnchor without ws: estimate from offsets only
    elif hasattr(anchor, "to") and anchor.to:
        to_col_off = getattr(anchor.to, "colOff", 0) or 0
        to_row_off = getattr(anchor.to, "rowOff", 0) or 0
        from_col_off = getattr(anchor._from, "colOff", 0) or 0
        from_row_off = getattr(anchor._from, "rowOff", 0) or 0

        # Rough estimate based on offsets and cell spans
        col_span = max(1, anchor.to.col - anchor._from.col)
        row_span = max(1, anchor.to.row - anchor._from.row)

        w = col_span * DEFAULT_COL_WIDTH + int(
            (to_col_off - from_col_off) / EMU_PER_PIXEL,
        )
        h = row_span * DEFAULT_ROW_HEIGHT + int(
            (to_row_off - from_row_off) / EMU_PER_PIXEL,
        )
        w = max(w, 10)
        h = max(h, 10)

    else:
        w, h = 100, 50  # Small default for unknown anchor types

    return col, row, w, h, int(w / 70) + 1, int(h / 20) + 1


# ============================================================================
# Image Rendering
# ============================================================================


def render_worksheet_images(
    img: Image.Image,
    ws,
    x_positions: dict,
    y_positions: dict,
    start_col_idx: int,
    start_row: int,
    end_col_idx: int,
    end_row: int,
    scale: float,
):
    """Render embedded images."""
    if not hasattr(ws, "_images") or not ws._images:
        return

    for image in ws._images:
        try:
            anchor = getattr(image, "anchor", None)
            if not anchor:
                continue

            col, row, w, h, _, _ = _get_anchor_extent(anchor, ws)
            if (
                col < start_col_idx
                or col > end_col_idx
                or row < start_row
                or row > end_row
            ):
                continue

            x = int(x_positions.get(col, 0) * scale)
            y = int(y_positions.get(row, 0) * scale)
            w, h = int(w * scale), int(h * scale)

            # Load image data
            img_data = None
            if hasattr(image, "ref") and image.ref:
                img_data = image.ref
            elif hasattr(image, "_data"):
                img_data = image._data()

            if img_data:
                img_pil = Image.open(
                    BytesIO(img_data) if isinstance(img_data, bytes) else img_data,
                )
                if img_pil.mode == "CMYK":
                    img_pil = img_pil.convert("RGB")
                img_pil = img_pil.resize(
                    (max(1, w), max(1, h)),
                    Image.Resampling.LANCZOS,
                )
                img.paste(img_pil, (x, y))
        except:
            continue


def render_worksheet_charts(
    img: Image.Image,
    draw: ImageDraw.Draw,
    ws,
    x_positions: dict,
    y_positions: dict,
    start_col_idx: int,
    start_row: int,
    end_col_idx: int,
    end_row: int,
    scale: float,
):
    """Render charts."""
    if not hasattr(ws, "_charts") or not ws._charts:
        return

    for chart in ws._charts:
        try:
            anchor = getattr(chart, "anchor", None)
            if not anchor:
                continue

            col, row, w, h, _, _ = _get_anchor_extent(anchor, ws)
            if (
                col < start_col_idx
                or col > end_col_idx + 1
                or row < start_row
                or row > end_row + 1
            ):
                continue

            x = int(x_positions.get(col, x_positions.get(start_col_idx, 0)) * scale)
            y = int(y_positions.get(row, y_positions.get(start_row, 0)) * scale)
            w, h = int(w * scale) or 300, int(h * scale) or 200

            chart_img = _render_chart_to_image(chart, ws, w, h)
            if chart_img:
                if chart_img.size != (w, h):
                    chart_img = chart_img.resize((w, h), Image.Resampling.LANCZOS)
                img.paste(chart_img, (x, y))
            else:
                # Placeholder
                draw.rectangle(
                    [x, y, x + w, y + h],
                    fill=(248, 248, 248),
                    outline=(180, 180, 180),
                )
                label = f"[{type(chart).__name__}]"
                font = get_default_font(9, scale=scale)
                bbox = draw.textbbox((0, 0), label, font=font)
                draw.text(
                    (x + (w - bbox[2]) // 2, y + (h - bbox[3]) // 2),
                    label,
                    fill=(120, 120, 120),
                    font=font,
                )
        except:
            continue


# ============================================================================
# Main Render Function
# ============================================================================


def render_excel_sheet(
    sheet: "Worksheet",
    cell_range: str | None = None,
    scale: float = 1.0,
) -> Image.Image:
    """
    Render an Excel sheet range as a PIL Image.

    Args:
        sheet: openpyxl Worksheet instance to render
        cell_range: Range like "A1:J20" (default: full used range including charts/images)
        scale: Scale factor for output

    Returns:
        PIL Image of the rendered range
    """
    ws = sheet
    wb = ws.parent
    theme_colors = extract_theme_colors(wb)

    # Determine range
    if cell_range is None:
        max_col, max_row = ws.max_column or 1, ws.max_row or 1

        # Expand for charts and images
        for obj_list in [getattr(ws, "_charts", []), getattr(ws, "_images", [])]:
            for obj in obj_list:
                try:
                    anchor = getattr(obj, "anchor", None)
                    if anchor:
                        col, row, _, _, ec, er = _get_anchor_extent(anchor, ws)
                        max_col = max(max_col, col + ec)
                        max_row = max(max_row, row + er)
                except:
                    pass

        cell_range = f"A1:{get_column_letter(max_col)}{max_row}"

    # Parse range and calculate dimensions
    start_col, start_row, end_col, end_row = parse_range(cell_range)
    start_col_idx, end_col_idx = column_index_from_string(
        start_col,
    ), column_index_from_string(end_col)

    col_widths = {
        i: get_column_width_pixels(ws, get_column_letter(i))
        for i in range(start_col_idx, end_col_idx + 1)
    }
    row_heights = {
        i: get_row_height_pixels(ws, i) for i in range(start_row, end_row + 1)
    }

    total_width = ROW_HEADER_WIDTH + sum(col_widths.values())
    total_height = COL_HEADER_HEIGHT + sum(row_heights.values())

    # Create image
    img = Image.new(
        "RGB",
        (int(total_width * scale), int(total_height * scale)),
        DEFAULT_CELL_BG,
    )
    draw = ImageDraw.Draw(img)

    # Build position maps
    x_positions, y_positions = {}, {}
    x = ROW_HEADER_WIDTH
    for col_idx in range(start_col_idx, end_col_idx + 2):
        x_positions[col_idx] = x
        x += col_widths.get(col_idx, 0)

    y = COL_HEADER_HEIGHT
    for row_idx in range(start_row, end_row + 2):
        y_positions[row_idx] = y
        y += row_heights.get(row_idx, 0)

    # Build merged cell map
    merged_map, merged_anchors = {}, set()
    for mr in ws.merged_cells.ranges:
        merged_anchors.add((mr.min_row, mr.min_col))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = mr

    header_font = get_default_font(DEFAULT_FONT_SIZE - 1, bold=True, scale=scale)

    # PASS 1: Cell backgrounds
    drawn_anchors = set()
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col_idx, end_col_idx + 1):
            key = (row_idx, col_idx)

            if key in merged_map:
                mr = merged_map[key]
                anchor = (mr.min_row, mr.min_col)

                # Skip if we've already drawn this merged cell's background
                if anchor in drawn_anchors:
                    continue

                # For merged cells, find the first visible cell (top-left within render range)
                # This handles cases where anchor is outside the rendered range
                first_visible_row = max(mr.min_row, start_row)
                first_visible_col = max(mr.min_col, start_col_idx)

                # Only draw when we're at the first visible cell of this merge
                if row_idx != first_visible_row or col_idx != first_visible_col:
                    continue

                drawn_anchors.add(anchor)

                x1 = int(x_positions[max(mr.min_col, start_col_idx)] * scale)
                y1 = int(y_positions[max(mr.min_row, start_row)] * scale)
                x2 = int(x_positions[min(mr.max_col, end_col_idx) + 1] * scale)
                y2 = int(y_positions[min(mr.max_row, end_row) + 1] * scale)

                cell = ws.cell(row=anchor[0], column=anchor[1])
                draw.rectangle(
                    [x1, y1, x2, y2],
                    fill=get_cell_background(cell, theme_colors),
                )
            else:
                x1, y1 = int(x_positions[col_idx] * scale), int(
                    y_positions[row_idx] * scale,
                )
                x2, y2 = int(x_positions[col_idx + 1] * scale), int(
                    y_positions[row_idx + 1] * scale,
                )
                cell = ws.cell(row=row_idx, column=col_idx)
                draw.rectangle(
                    [x1, y1, x2, y2],
                    fill=get_cell_background(cell, theme_colors),
                )

    # PASS 2: Grid lines (skip inside merged cells)
    for row_idx in range(start_row, end_row + 2):
        y = int(y_positions.get(row_idx, y_positions[end_row + 1]) * scale)
        col_idx = start_col_idx
        while col_idx <= end_col_idx:
            x_start = int(x_positions[col_idx] * scale)

            # Check if line should be skipped (inside merged cell)
            skip = False
            if row_idx > start_row:
                above = (row_idx - 1, col_idx)
                if above in merged_map:
                    mr = merged_map[above]
                    if mr.min_row <= row_idx - 1 < mr.max_row:
                        skip = True
                        col_idx = mr.max_col + 1
                        continue

            # Find end of line segment
            end_col = col_idx
            while end_col <= end_col_idx:
                next_key = (row_idx - 1, end_col) if row_idx > start_row else None
                if next_key and next_key in merged_map:
                    mr = merged_map[next_key]
                    if mr.min_row <= row_idx - 1 < mr.max_row:
                        break
                end_col += 1

            x_end = int(x_positions.get(end_col, x_positions[end_col_idx + 1]) * scale)
            draw.line([(x_start, y), (x_end, y)], fill=GRID_COLOR, width=1)
            col_idx = end_col

    for col_idx in range(start_col_idx, end_col_idx + 2):
        x = int(x_positions.get(col_idx, x_positions[end_col_idx + 1]) * scale)
        row_idx = start_row
        while row_idx <= end_row:
            y_start = int(y_positions[row_idx] * scale)

            # Check for merged cell
            if col_idx > start_col_idx:
                left = (row_idx, col_idx - 1)
                if left in merged_map:
                    mr = merged_map[left]
                    if mr.min_col <= col_idx - 1 < mr.max_col:
                        row_idx = mr.max_row + 1
                        continue

            # Find end of segment
            end_row_seg = row_idx
            while end_row_seg <= end_row:
                next_key = (
                    (end_row_seg, col_idx - 1) if col_idx > start_col_idx else None
                )
                if next_key and next_key in merged_map:
                    mr = merged_map[next_key]
                    if mr.min_col <= col_idx - 1 < mr.max_col:
                        break
                end_row_seg += 1

            y_end = int(y_positions.get(end_row_seg, y_positions[end_row + 1]) * scale)
            draw.line([(x, y_start), (x, y_end)], fill=GRID_COLOR, width=1)
            row_idx = end_row_seg

    # PASS 3: Cell borders
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue

            x1, y1 = int(x_positions[col_idx] * scale), int(
                y_positions[row_idx] * scale,
            )

            # Get cell extent (handle merged)
            if (row_idx, col_idx) in merged_map:
                mr = merged_map[(row_idx, col_idx)]
                x2 = int(x_positions[min(mr.max_col, end_col_idx) + 1] * scale)
                y2 = int(y_positions[min(mr.max_row, end_row) + 1] * scale)
            else:
                x2 = int(x_positions[col_idx + 1] * scale)
                y2 = int(y_positions[row_idx + 1] * scale)

            draw_cell_borders(draw, cell, x1, y1, x2, y2, theme_colors, scale)

    # PASS 4: Cell text
    drawn_text = set()
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col_idx, end_col_idx + 1):
            key = (row_idx, col_idx)

            # Handle merged cells
            if key in merged_map:
                mr = merged_map[key]
                anchor = (mr.min_row, mr.min_col)

                # Skip if we've already drawn this merged cell
                if anchor in drawn_text:
                    continue

                # For merged cells, find the first visible cell (top-left within render range)
                # This handles cases where anchor is outside the rendered range
                first_visible_row = max(mr.min_row, start_row)
                first_visible_col = max(mr.min_col, start_col_idx)

                # Only draw when we're at the first visible cell of this merge
                if row_idx != first_visible_row or col_idx != first_visible_col:
                    continue

                drawn_text.add(anchor)
                cell = ws.cell(row=anchor[0], column=anchor[1])

                x1 = int(x_positions[max(mr.min_col, start_col_idx)] * scale)
                y1 = int(y_positions[max(mr.min_row, start_row)] * scale)
                x2 = int(x_positions[min(mr.max_col, end_col_idx) + 1] * scale)
                y2 = int(y_positions[min(mr.max_row, end_row) + 1] * scale)
            else:
                cell = ws.cell(row=row_idx, column=col_idx)
                x1, y1 = int(x_positions[col_idx] * scale), int(
                    y_positions[row_idx] * scale,
                )
                x2, y2 = int(x_positions[col_idx + 1] * scale), int(
                    y_positions[row_idx + 1] * scale,
                )

            if cell.value is None:
                continue

            text, color_override = format_value_with_color(cell)
            if not text:
                continue

            # Use symbol font for special Unicode characters, otherwise regular font
            if has_special_symbols(text):
                font = get_symbol_font(
                    (
                        cell.font.size
                        if hasattr(cell, "font") and cell.font and cell.font.size
                        else DEFAULT_FONT_SIZE
                    ),
                    scale,
                )
            else:
                font = get_font_for_cell(cell, scale)

            color = color_override or get_text_color(cell, theme_colors)
            scaled_padding = int(CELL_PADDING * scale)
            cw, ch = x2 - x1, y2 - y1

            # Truncate text to fit within cell width (with padding on both sides)
            max_text_width = cw - (2 * scaled_padding)
            text = truncate_text_to_width(draw, text, font, max_text_width)
            if not text:
                continue

            # Calculate text position
            bbox = draw.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]

            # Horizontal alignment
            align = getattr(cell, "alignment", None)
            h_align = align.horizontal if align and align.horizontal else None
            if h_align == "center":
                tx = x1 + (cw - tw) // 2
            elif h_align == "right":
                tx = x2 - tw - scaled_padding
            else:
                tx = x1 + scaled_padding

            # Vertical alignment
            v_align = align.vertical if align and align.vertical else None
            if v_align == "center":
                ty = y1 + (ch - th) // 2
            elif v_align == "bottom":
                ty = y2 - th - scaled_padding
            else:
                ty = y1 + scaled_padding

            draw.text((tx, ty), text, fill=color, font=font)

    # PASS 5: Headers with grid lines
    scaled_header_h = int(COL_HEADER_HEIGHT * scale)
    scaled_header_w = int(ROW_HEADER_WIDTH * scale)

    # Column header background
    draw.rectangle(
        [0, 0, int(total_width * scale), scaled_header_h],
        fill=HEADER_BG_COLOR,
    )

    # Draw vertical grid lines through column header
    for col_idx in range(start_col_idx, end_col_idx + 2):
        x = int(x_positions[col_idx] * scale)
        draw.line([(x, 0), (x, scaled_header_h)], fill=HEADER_BORDER_COLOR, width=1)

    # Column header text
    for col_idx in range(start_col_idx, end_col_idx + 1):
        x1 = int(x_positions[col_idx] * scale)
        x2 = int(x_positions[col_idx + 1] * scale)
        letter = get_column_letter(col_idx)
        bbox = draw.textbbox((0, 0), letter, font=header_font)
        tx = x1 + (x2 - x1 - bbox[2]) // 2
        ty = int((scaled_header_h - bbox[3]) // 2)
        draw.text((tx, ty), letter, fill=(0, 0, 0), font=header_font)

    # Row header background
    draw.rectangle(
        [0, 0, scaled_header_w, int(total_height * scale)],
        fill=HEADER_BG_COLOR,
    )

    # Draw horizontal grid lines through row header
    for row_idx in range(start_row, end_row + 2):
        y = int(y_positions[row_idx] * scale)
        draw.line([(0, y), (scaled_header_w, y)], fill=HEADER_BORDER_COLOR, width=1)

    # Row header text
    for row_idx in range(start_row, end_row + 1):
        y1 = int(y_positions[row_idx] * scale)
        y2 = int(y_positions[row_idx + 1] * scale)
        text = str(row_idx)
        bbox = draw.textbbox((0, 0), text, font=header_font)
        tx = int((scaled_header_w - bbox[2]) // 2)
        ty = y1 + (y2 - y1 - bbox[3]) // 2
        draw.text((tx, ty), text, fill=(0, 0, 0), font=header_font)

    # Main header borders (separating headers from cells)
    draw.line(
        [(scaled_header_w, 0), (scaled_header_w, int(total_height * scale))],
        fill=HEADER_BORDER_COLOR,
        width=1,
    )
    draw.line(
        [(0, scaled_header_h), (int(total_width * scale), scaled_header_h)],
        fill=HEADER_BORDER_COLOR,
        width=1,
    )

    # PASS 6: Images and charts
    render_worksheet_images(
        img,
        ws,
        x_positions,
        y_positions,
        start_col_idx,
        start_row,
        end_col_idx,
        end_row,
        scale,
    )
    render_worksheet_charts(
        img,
        draw,
        ws,
        x_positions,
        y_positions,
        start_col_idx,
        start_row,
        end_col_idx,
        end_row,
        scale,
    )

    return img


# ============================================================================
# PDF Rendering
# ============================================================================


def render_pdf(
    source: str | Path | "pymupdf.Document",
    page: int = 0,
    dpi: int = 150,
) -> Image.Image:
    """
    Render a PDF page as a PIL Image.

    Args:
        source: Path to PDF file or pymupdf Document object
        page: Page number to render (0-indexed, default: 0)
        dpi: Resolution for rendering (default: 150)

    Returns:
        PIL Image of the rendered PDF page

    Example:
        >>> img = render_pdf("document.pdf")  # First page
        >>> img = render_pdf("document.pdf", page=2)  # Third page
        >>> import pymupdf
        >>> doc = pymupdf.open("document.pdf")
        >>> img = render_pdf(doc, page=1)
    """
    import pymupdf

    # Handle source type
    should_close = isinstance(source, (str, Path))
    doc = pymupdf.open(source) if should_close else source

    try:
        # Validate page number
        if page < 0 or page >= len(doc):
            raise ValueError(
                f"Page {page} out of range (document has {len(doc)} pages)",
            )

        # Get the page and render to pixmap
        pdf_page = doc[page]

        # Calculate zoom factor from DPI (72 is PDF default DPI)
        zoom = dpi / 72
        matrix = pymupdf.Matrix(zoom, zoom)

        # Render page to pixmap
        pixmap = pdf_page.get_pixmap(matrix=matrix, alpha=False)

        # Convert to PIL Image
        img = Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)

        return img
    finally:
        if should_close:
            doc.close()
