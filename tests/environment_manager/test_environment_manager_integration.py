"""Integration test for EnvironmentManager with a realistic examplecorp-like environment.

This test uploads a multi-file environment that:
- Has a service class with real domain logic (Excel generation via openpyxl)
- Has a Pydantic schemas module imported cross-file
- Has a BaseEnvironment subclass tying them together
- Declares openpyxl as a dependency

It validates the full upload -> load -> call flow, including that the
loaded service actually produces a valid .xlsx file.
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import pytest

from unity.environment_manager.simulated import SimulatedEnvironmentManager
from unity.actor.environments.base import BaseEnvironment

# ── Source files for the "examplecorp-like" environment ──────────────────────────

SCHEMAS_SOURCE = '''\
"""Pydantic schemas for financial data extraction."""
from __future__ import annotations

from typing import Optional
from pydantic import BaseModel, Field


class FieldValue(BaseModel):
    value: float | int | str | None = None
    source: str | None = None


class IncomeData(BaseModel):
    fee_income: FieldValue | None = None
    other_income: FieldValue | None = None
    total_income: FieldValue | None = None


class ExpenditureData(BaseModel):
    wages: FieldValue | None = None
    provisions: FieldValue | None = None
    total_expenditure: FieldValue | None = None


class FiscalYearData(BaseModel):
    property_name: str
    fiscal_year: str
    income: IncomeData | None = None
    expenditure: ExpenditureData | None = None


FINANCIAL_DATA_SCHEMA = {
    "INCOME": ["Fee Income", "Other Income", "Total Income"],
    "EXPENDITURE": ["Wages", "Provisions", "Total Expenditure"],
}

FIELD_NAME_MAPPING = {
    "Fee Income": "fee_income",
    "Other Income": "other_income",
    "Total Income": "total_income",
    "Wages": "wages",
    "Provisions": "provisions",
    "Total Expenditure": "total_expenditure",
}

CATEGORY_TO_SECTION = {
    "INCOME": "income",
    "EXPENDITURE": "expenditure",
}
'''


SERVICE_SOURCE = '''\
"""examplecorp-like service that generates Excel spreadsheets from financial data."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from schemas import (
    CATEGORY_TO_SECTION,
    FIELD_NAME_MAPPING,
    FINANCIAL_DATA_SCHEMA,
    FiscalYearData,
)


class FinancialService:
    """Domain tools for financial deliverables."""

    def create_financial_excel(
        self,
        json_file_path: str,
        output_path: str,
    ) -> str:
        """Read validated financial JSON and produce a formatted Excel spreadsheet.

        Parameters
        ----------
        json_file_path : str
            Path to a JSON file containing ``list[FiscalYearData]``.
        output_path : str
            Where to write the ``.xlsx`` output.

        Returns
        -------
        str
            The path to the created Excel file.
        """
        raw = json.loads(Path(json_file_path).read_text(encoding="utf-8"))
        fiscal_data = [FiscalYearData.model_validate(item) for item in raw]

        properties: dict[str, list[FiscalYearData]] = defaultdict(list)
        for item in fiscal_data:
            properties[item.property_name].append(item)
        for prop_name in properties:
            properties[prop_name].sort(key=lambda x: x.fiscal_year)

        wb = Workbook()
        wb.remove(wb.active)

        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid",
        )
        header_font = Font(bold=True, size=10, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for prop_name, fiscal_years in properties.items():
            ws = wb.create_sheet(title=prop_name[:31])
            num_years = len(fiscal_years)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_years + 1)
            ws.cell(row=1, column=1, value="HISTORIC ACCOUNTS").font = title_font

            for col_idx, fy in enumerate(fiscal_years, start=2):
                cell = ws.cell(row=3, column=col_idx, value=f"YE {fy.fiscal_year}")
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            current_row = 5
            for category, fields in FINANCIAL_DATA_SCHEMA.items():
                ws.cell(row=current_row, column=1, value=category).font = Font(bold=True)
                current_row += 1
                for field_name in fields:
                    ws.cell(row=current_row, column=1, value=field_name)
                    pydantic_field = FIELD_NAME_MAPPING.get(field_name)
                    section_name = CATEGORY_TO_SECTION[category]
                    for col_idx, fy in enumerate(fiscal_years, start=2):
                        section_data = getattr(fy, section_name, None)
                        value = None
                        if section_data is not None:
                            field_obj = getattr(section_data, pydantic_field, None)
                            if field_obj is not None and hasattr(field_obj, "value"):
                                value = field_obj.value
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = thin_border
                    current_row += 1
                current_row += 1

        out = Path(output_path)
        wb.save(out)
        return str(out)
'''


ENV_SOURCE = '''\
"""BaseEnvironment subclass for the financial service."""
from __future__ import annotations

import json
from typing import Any, Dict

from unity.actor.environments.base import BaseEnvironment, ToolMetadata
from schemas import FiscalYearData
from service import FinancialService


class FinancialEnvironment(BaseEnvironment):
    """Financial data extraction environment.

    Namespace: ``financial``
    """

    NAMESPACE = "financial"

    def __init__(self) -> None:
        self._service = FinancialService()
        super().__init__(instance=self._service, namespace=self.NAMESPACE)

    def get_tools(self) -> Dict[str, ToolMetadata]:
        return {
            f"{self.NAMESPACE}.create_financial_excel": ToolMetadata(
                name=f"{self.NAMESPACE}.create_financial_excel",
                is_impure=True,
            ),
        }

    def get_prompt_context(self) -> str:
        schema = json.dumps(FiscalYearData.model_json_schema(), indent=2)
        return (
            f"### `{self.NAMESPACE}` — Financial Tools\\n\\n"
            f"**`{self.NAMESPACE}.create_financial_excel(json_file_path, output_path)`**\\n"
            "  Produce a formatted HISTORIC ACCOUNTS Excel spreadsheet.\\n\\n"
            "#### FiscalYearData Schema\\n\\n"
            f"```json\\n{schema}\\n```\\n"
        )

    async def capture_state(self) -> Dict[str, Any]:
        return {"type": "financial"}


financial_env = FinancialEnvironment()
'''


# ── Test data ─────────────────────────────────────────────────────────────────

SAMPLE_FINANCIAL_DATA = [
    {
        "property_name": "Sunrise Care Home",
        "fiscal_year": "2023",
        "income": {
            "fee_income": {"value": 1500000, "source": "P&L page 2"},
            "other_income": {"value": 50000, "source": "P&L page 2"},
            "total_income": {"value": 1550000, "source": "P&L page 2"},
        },
        "expenditure": {
            "wages": {"value": 800000, "source": "P&L page 3"},
            "provisions": {"value": 120000, "source": "P&L page 3"},
            "total_expenditure": {"value": 920000, "source": "P&L page 3"},
        },
    },
    {
        "property_name": "Sunrise Care Home",
        "fiscal_year": "2024",
        "income": {
            "fee_income": {"value": 1600000, "source": "P&L page 2"},
            "other_income": {"value": 55000, "source": "P&L page 2"},
            "total_income": {"value": 1655000, "source": "P&L page 2"},
        },
        "expenditure": {
            "wages": {"value": 850000, "source": "P&L page 3"},
            "provisions": {"value": 130000, "source": "P&L page 3"},
            "total_expenditure": {"value": 980000, "source": "P&L page 3"},
        },
    },
]


# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture
def manager():
    return SimulatedEnvironmentManager()


# ── Integration tests ─────────────────────────────────────────────────────────


class TestCollierLikeIntegration:
    """End-to-end tests with a realistic multi-file environment using openpyxl."""

    def _upload_financial_env(self, manager):
        return manager.upload_environment(
            name="financial",
            files={
                "schemas.py": SCHEMAS_SOURCE,
                "service.py": SERVICE_SOURCE,
                "financial_env.py": ENV_SOURCE,
            },
            dependencies=["openpyxl"],
            env="financial_env:financial_env",
        )

    def test_upload_and_list(self, manager):
        eid = self._upload_financial_env(manager)
        envs = manager.list_environments()
        assert len(envs) == 1
        assert envs[0]["name"] == "financial"
        assert envs[0]["environment_id"] == eid

    def test_load_returns_base_environment(self, manager):
        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        assert isinstance(env, BaseEnvironment)

    def test_namespace(self, manager):
        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        assert env.namespace == "financial"

    def test_tools(self, manager):
        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        tools = env.get_tools()
        assert "financial.create_financial_excel" in tools
        assert tools["financial.create_financial_excel"].is_impure is True

    def test_prompt_context_includes_schema(self, manager):
        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        ctx = env.get_prompt_context()
        assert "financial" in ctx
        assert "create_financial_excel" in ctx
        assert "FiscalYearData" in ctx

    def test_service_produces_excel(self, manager):
        """The crown jewel: upload, load, call the service, get a real .xlsx."""
        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        service = env.get_instance()

        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "data.json"
            json_path.write_text(json.dumps(SAMPLE_FINANCIAL_DATA))

            output_path = Path(tmpdir) / "output.xlsx"
            result = service.create_financial_excel(
                str(json_path),
                str(output_path),
            )

            assert Path(result).exists()
            assert Path(result).suffix == ".xlsx"
            assert Path(result).stat().st_size > 0

    def test_excel_content(self, manager):
        """Verify the generated Excel has the right structure."""
        from openpyxl import load_workbook

        eid = self._upload_financial_env(manager)
        env = manager.load_environment(eid)
        service = env.get_instance()

        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "data.json"
            json_path.write_text(json.dumps(SAMPLE_FINANCIAL_DATA))

            output_path = Path(tmpdir) / "output.xlsx"
            service.create_financial_excel(str(json_path), str(output_path))

            wb = load_workbook(str(output_path))
            assert "Sunrise Care Home" in wb.sheetnames
            ws = wb["Sunrise Care Home"]
            assert ws.cell(row=1, column=1).value == "HISTORIC ACCOUNTS"
            # Check fiscal year headers
            assert "YE 2023" in ws.cell(row=3, column=2).value
            assert "YE 2024" in ws.cell(row=3, column=3).value

    def test_load_all_environments(self, manager):
        self._upload_financial_env(manager)
        envs = manager.load_all_environments()
        assert len(envs) == 1
        assert envs[0].namespace == "financial"

    def test_delete_and_reload(self, manager):
        eid = self._upload_financial_env(manager)
        manager.delete_environment(environment_id=eid)
        assert manager.list_environments() == []
        with pytest.raises(ValueError):
            manager.load_environment(eid)
