"""
Comprehensive tests for parsing different file formats with DoclingParser.
Tests all supported formats (.txt, .pdf, .docx) with rigorous validation.
"""

from __future__ import annotations

import re
import time

import pytest

from tests.helpers import _handle_project
from tests.test_file_manager.conftest import llm_judge_html_equivalence


def normalize_html_for_comparison(html_str: str) -> str:
    """
    Normalize HTML for comparison by removing ALL attributes and standardizing structure.

    This removes:
    - ALL HTML attributes from ALL tags (everything between <tag and >)
    - <thead> wrappers (pandas adds these, Docling might not)
    - Extra whitespace

    Returns lowercase normalized HTML string with pure tags only.
    """
    # First, normalize whitespace
    html_str = re.sub(r"\s+", " ", html_str.strip())
    html_str = re.sub(r">\s+<", "><", html_str)

    # Remove ALL attributes from opening tags
    # This handles: <tag attr="value">, <tag attr='value'>, <tag attr>, etc.
    # Pattern explanation:
    # <        - opening bracket
    # (\w+)    - capture the tag name (group 1)
    # \s+      - at least one whitespace (means there are attributes)
    # [^>]*    - anything that's not a closing bracket (the attributes)
    # >        - closing bracket
    # Replace with: <tagname>
    html_str = re.sub(r"<(\w+)\s+[^>]*>", r"<\1>", html_str)

    # Also handle self-closing tags like <br /> or <img ... />
    html_str = re.sub(r"<(\w+)\s+[^/>]*/>", r"<\1/>", html_str)

    # Normalize thead/tbody structure - remove thead wrapper if present
    html_str = re.sub(r"<thead>(.*?)</thead>", r"\1", html_str)

    return html_str.lower()


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_txt_simple(parser, supported_format_files):
    """Test parsing simple text file."""
    txt_files = supported_format_files[".txt"]["files"]
    txt_file = txt_files["simple"]
    doc = parser.parse(txt_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/plain"
    assert doc.metadata.file_format.value == "txt"
    assert doc.metadata.file_name.endswith(".txt")

    # Check content is preserved
    full_text = doc.to_plain_text()
    assert "simple text file" in full_text.lower()

    # Should have basic structure
    assert len(doc.sections) >= 0  # May have 0 or more sections
    assert doc.processing_status == "completed"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_txt_multi_paragraph(parser, supported_format_files):
    """Test parsing multi-paragraph text file."""
    txt_files = supported_format_files[".txt"]["files"]
    txt_file = txt_files["multi_paragraph"]
    doc = parser.parse(txt_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/plain"
    assert doc.metadata.file_format.value == "txt"

    # Check content preservation
    full_text = doc.to_plain_text()
    assert "First paragraph" in full_text
    assert "Second paragraph" in full_text
    assert "Third paragraph" in full_text

    # Should have structure
    assert len(doc.sections) >= 1
    assert doc.processing_status == "completed"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_txt_special_characters(parser, supported_format_files):
    """Test parsing text file with special characters."""
    txt_files = supported_format_files[".txt"]["files"]
    txt_file = txt_files["special_chars"]
    doc = parser.parse(txt_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/plain"
    assert doc.metadata.file_format.value == "txt"

    # Check Unicode handling
    full_text = doc.to_plain_text()
    # Check for at least some special characters from our fixture
    has_special_chars = any(
        char in full_text for char in ["café", "naïve", "€", "你好"]
    )
    assert has_special_chars, f"Expected special characters in: {full_text}"

    assert doc.processing_status == "completed"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_pdf(parser):
    """Test parsing PDF file from sample directory."""
    from pathlib import Path

    # Use the actual PDF file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    pdf_file = sample_dir / "IT_Department_Policy_Document.pdf"

    if not pdf_file.exists():
        pytest.skip("PDF sample file not found")

    doc = parser.parse(pdf_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "application/pdf"
    assert doc.metadata.file_format.value == "pdf"
    assert doc.metadata.file_name == "IT_Department_Policy_Document.pdf"
    assert doc.metadata.file_size > 0

    # Check content extraction
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "PDF should contain extractable text"

    # Should have some structure (PDFs typically have multiple sections)
    assert len(doc.sections) >= 1
    assert doc.processing_status == "completed"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_docx(parser):
    """Test parsing DOCX file from sample directory."""
    from pathlib import Path

    # Use the actual DOCX file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    docx_file = sample_dir / "SmartHome_Hub_X200_Technical_Documentation.docx"

    if not docx_file.exists():
        pytest.skip("DOCX sample file not found")

    doc = parser.parse(docx_file)

    # Check metadata
    assert doc.metadata.mime_type.value == (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    assert doc.metadata.file_format.value == "docx"
    assert doc.metadata.file_name == "SmartHome_Hub_X200_Technical_Documentation.docx"
    assert doc.metadata.file_size > 0

    # Check content extraction
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "DOCX should contain extractable text"

    # DOCX files typically have good structure
    assert len(doc.sections) >= 1
    assert doc.processing_status == "completed"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv(parser):
    """Test parsing CSV file from sample directory."""
    from pathlib import Path

    # Use the actual CSV file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    csv_file = sample_dir / "employee_records.csv"

    if not csv_file.exists():
        pytest.skip("CSV sample file not found")

    doc = parser.parse(csv_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/csv"
    assert doc.metadata.file_format.value == "csv"
    assert doc.metadata.file_name == "employee_records.csv"
    assert doc.metadata.file_size > 0
    assert doc.processing_status == "completed"

    # Check content extraction - verify key data from CSV
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "CSV should contain extractable text"

    # Verify specific employee records are present
    assert "Alice Johnson" in full_text, "Should contain Alice Johnson record"
    assert "Bob Smith" in full_text, "Should contain Bob Smith record"
    assert "Charlie Davis" in full_text, "Should contain Charlie Davis record"
    assert "Diana Green" in full_text, "Should contain Diana Green record"
    assert "Ethan Brown" in full_text, "Should contain Ethan Brown record"
    assert "Fiona White" in full_text, "Should contain Fiona White record"

    # Verify departments are present
    assert "Engineering" in full_text, "Should contain Engineering department"
    assert "Marketing" in full_text, "Should contain Marketing department"
    assert "Sales" in full_text, "Should contain Sales department"
    assert "Finance" in full_text, "Should contain Finance department"
    assert "HR" in full_text, "Should contain HR department"

    # Verify salary data is present (at least some values)
    assert (
        "85000" in full_text or "85,000" in full_text
    ), "Should contain Alice's salary"
    assert "72000" in full_text or "72,000" in full_text, "Should contain Bob's salary"

    # Verify column headers are preserved
    assert (
        "EmployeeID" in full_text or "Employee" in full_text
    ), "Should contain EmployeeID header"
    assert "Name" in full_text, "Should contain Name header"
    assert "Department" in full_text, "Should contain Department header"
    assert "Salary" in full_text, "Should contain Salary header"

    # Check structure - CSV produces tables, not sections
    assert len(doc.metadata.tables) >= 1, "CSV should produce at least one table"

    # Check statistics
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0
    assert doc.metadata.total_sections >= 0

    # Validate table extraction with pandas HTML comparison (if pandas available)
    try:
        import pandas as pd

        # Check that tables were extracted
        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        # Load the CSV with pandas for ground truth comparison
        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        # Get the extracted table HTML from Docling
        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        # Normalize both HTML strings for comparison
        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        # If pandas is not available, we still pass the test since core assertions passed
        print(f"Skipping pandas verification for test_csv")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_xlsx(parser):
    """Test parsing XLSX file from sample directory."""
    from pathlib import Path

    # Use the actual XLSX file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    xlsx_file = sample_dir / "project_status.xlsx"

    if not xlsx_file.exists():
        pytest.skip("XLSX sample file not found")

    doc = parser.parse(xlsx_file)

    # Check metadata
    mt = doc.metadata.mime_type.value.lower()
    assert (
        "xlsx" in mt or "spreadsheet" in mt or "openxmlformats" in mt
    ), f"Expected XLSX file type, got: {doc.metadata.mime_type}"
    assert doc.metadata.file_format.value == "xlsx"
    assert doc.metadata.file_name == "project_status.xlsx"
    assert doc.metadata.file_size > 0
    assert doc.processing_status == "completed"

    # Check content extraction - verify key data from XLSX
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "XLSX should contain extractable text"

    # Verify project IDs are present (PRJ-001 through PRJ-005)
    assert (
        "PRJ-001" in full_text or "PRJ 001" in full_text
    ), "Should contain project PRJ-001"
    assert (
        "PRJ-002" in full_text or "PRJ 002" in full_text
    ), "Should contain project PRJ-002"
    assert (
        "PRJ-003" in full_text or "PRJ 003" in full_text
    ), "Should contain project PRJ-003"
    assert (
        "PRJ-004" in full_text or "PRJ 004" in full_text
    ), "Should contain project PRJ-004"
    assert (
        "PRJ-005" in full_text or "PRJ 005" in full_text
    ), "Should contain project PRJ-005"

    # Verify project names are present
    assert (
        "AI Chatbot" in full_text or "Chatbot" in full_text
    ), "Should contain AI Chatbot project"
    assert "Website" in full_text, "Should contain Website project"
    assert "Mobile" in full_text, "Should contain Mobile project"
    assert (
        "Data Migration" in full_text or "Migration" in full_text
    ), "Should contain Data Migration project"
    assert "Security" in full_text, "Should contain Security project"

    # Verify manager names
    assert (
        "Alice" in full_text or "Johnson" in full_text
    ), "Should contain Alice Johnson as manager"
    assert (
        "Bob" in full_text or "Smith" in full_text
    ), "Should contain Bob Smith as manager"
    assert (
        "Charlie" in full_text or "Davis" in full_text
    ), "Should contain Charlie Davis as manager"
    assert (
        "Diana" in full_text or "Green" in full_text
    ), "Should contain Diana Green as manager"
    assert (
        "Ethan" in full_text or "Brown" in full_text
    ), "Should contain Ethan Brown as manager"

    # Verify project statuses
    assert "Completed" in full_text, "Should contain Completed status"
    assert (
        "In Progress" in full_text or "Progress" in full_text
    ), "Should contain In Progress status"
    assert "Delayed" in full_text, "Should contain Delayed status"
    assert "Planning" in full_text, "Should contain Planning status"

    # Verify budget data (at least some values)
    assert (
        "50000" in full_text or "50,000" in full_text
    ), "Should contain budget value 50000"
    assert (
        "75000" in full_text or "75,000" in full_text
    ), "Should contain budget value 75000"
    assert (
        "120000" in full_text or "120,000" in full_text
    ), "Should contain budget value 120000"

    # Verify column headers
    assert "Project" in full_text, "Should contain Project-related header"
    assert "Manager" in full_text, "Should contain Manager header"
    assert "Status" in full_text, "Should contain Status header"
    assert "Budget" in full_text, "Should contain Budget header"

    # Check structure - XLSX produces tables, not sections
    assert len(doc.metadata.tables) >= 1, "XLSX should produce at least one table"

    # Check statistics
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0
    assert doc.metadata.total_sections >= 0

    # Validate table extraction with pandas HTML comparison (if pandas available)
    try:
        import pandas as pd

        # Check that tables were extracted
        assert len(doc.metadata.tables) > 0, "Should extract table metadata from XLSX"

        # Load the XLSX with pandas for ground truth comparison
        df_expected = pd.read_excel(xlsx_file)
        expected_html = df_expected.to_html(index=False)

        # Get the extracted table HTML from Docling
        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        # Normalize both HTML strings for comparison
        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        # If pandas is not available, we still pass the test since core assertions passed
        print(f"Skipping pandas verification for test_xlsx")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_empty_txt(parser, supported_format_files):
    """Test parsing an empty text file."""
    txt_files = supported_format_files[".txt"]["files"]
    empty_file = txt_files["empty"]
    doc = parser.parse(empty_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/plain"
    assert doc.processing_status == "completed"

    # Empty file should have empty content
    full_text = doc.to_plain_text()
    assert full_text.strip() == ""

    # May have 0 sections for empty files
    assert len(doc.sections) >= 0


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_formats_comprehensive(parser):
    """Test parsing all supported formats comprehensively."""
    from pathlib import Path

    # Test that parser only supports the expected formats
    supported_formats = parser.supported_formats
    expected_formats = {".txt", ".pdf", ".docx"}

    # Verify all expected formats are supported
    for fmt in expected_formats:
        assert fmt in supported_formats, f"Expected format {fmt} to be supported"

    # Test each supported format with sample files
    sample_dir = Path(__file__).parent.parent / "sample"

    # Test PDF if available
    pdf_file = sample_dir / "IT_Department_Policy_Document.pdf"
    if pdf_file.exists():
        pdf_doc = parser.parse(pdf_file)
        assert pdf_doc.metadata.mime_type.value == "application/pdf"
        assert pdf_doc.metadata.file_format.value == "pdf"
        assert pdf_doc.processing_status == "completed"
        assert len(pdf_doc.to_plain_text().strip()) > 0

    # Test DOCX if available
    docx_file = sample_dir / "SmartHome_Hub_X200_Technical_Documentation.docx"
    if docx_file.exists():
        docx_doc = parser.parse(docx_file)
        assert docx_doc.metadata.mime_type.value == (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        assert docx_doc.metadata.file_format.value == "docx"
        assert docx_doc.processing_status == "completed"
        assert len(docx_doc.to_plain_text().strip()) > 0


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_binary_metadata(parser):
    """Test that binary formats (PDF, DOCX) have correct metadata."""
    from pathlib import Path

    sample_dir = Path(__file__).parent.parent / "sample"

    # Test PDF metadata
    pdf_file = sample_dir / "IT_Department_Policy_Document.pdf"
    if pdf_file.exists():
        pdf_doc = parser.parse(pdf_file)

        # Check all required metadata fields
        assert pdf_doc.metadata.file_name == "IT_Department_Policy_Document.pdf"
        assert pdf_doc.metadata.mime_type.value == "application/pdf"
        assert pdf_doc.metadata.file_format.value == "pdf"
        assert pdf_doc.metadata.file_size > 0
        assert pdf_doc.metadata.parser_name == "DoclingParser"
        assert pdf_doc.metadata.parser_version == "1.0.0"
        assert pdf_doc.metadata.created_at is not None
        assert pdf_doc.metadata.processed_at is not None

    # Test DOCX metadata
    docx_file = sample_dir / "SmartHome_Hub_X200_Technical_Documentation.docx"
    if docx_file.exists():
        docx_doc = parser.parse(docx_file)

        # Check all required metadata fields
        assert (
            docx_doc.metadata.file_name
            == "SmartHome_Hub_X200_Technical_Documentation.docx"
        )
        assert docx_doc.metadata.mime_type.value == (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        assert docx_doc.metadata.file_format.value == "docx"
        assert docx_doc.metadata.file_size > 0
        assert docx_doc.metadata.parser_name == "DoclingParser"
        assert docx_doc.metadata.parser_version == "1.0.0"
        assert docx_doc.metadata.created_at is not None
        assert docx_doc.metadata.processed_at is not None


# =============================================================================
# COMPREHENSIVE DYNAMIC FORMAT TESTS
# =============================================================================


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_all_supported_formats_dynamic(
    parser,
    supported_format_files,
    parser_validation_suite,
):
    """Test dynamic parsing of all supported formats with rigorous validation."""
    validation = parser_validation_suite

    for fmt, format_info in supported_format_files.items():
        print(f"\nTesting format: {fmt}")

        # Test the simple variant for each format
        if format_info.get("is_binary", False):
            # For binary formats, use sample files
            sample_files = [
                f
                for f in format_info["files"].values()
                if f.name.startswith("file_example")
            ]
            if sample_files:
                file_path = sample_files[0]  # Use first sample file
                print(f"Testing binary file: {file_path}")

                doc = parser.parse(file_path)

                # Validate basic structure
                assert doc.metadata.mime_type.value == format_info["mime_type"]
                validation["validate_structure"](
                    doc,
                    format_info["structure_expectations"],
                )

                # For binary files, just ensure we got some content
                assert (
                    doc.to_plain_text().strip()
                ), f"Binary file {fmt} should extract some text content"
        else:
            # For text formats, test the simple variant
            simple_file = format_info["files"]["simple"]
            print(f"Testing text file: {simple_file}")

            doc = parser.parse(simple_file)

            # Validate metadata
            assert doc.metadata.mime_type.value == format_info["mime_type"]

            # CSV/XLSX produce tables, not sections – adjust validation
            if fmt in (".csv", ".xlsx"):
                assert len(doc.metadata.tables) >= 1
                if format_info["validation_patterns"]:
                    validation["validate_content"](
                        doc,
                        format_info["validation_patterns"],
                    )
            else:
                # Validate structure meets expectations
                validation["validate_structure"](
                    doc,
                    format_info["structure_expectations"],
                )
                # Validate content preservation
                if format_info["validation_patterns"]:
                    validation["validate_content"](
                        doc,
                        format_info["validation_patterns"],
                    )


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_empty_files_all_formats(parser, supported_format_files):
    """Test handling of empty files across all supported formats."""

    for fmt, format_info in supported_format_files.items():
        if format_info.get("is_binary", False):
            continue  # Skip binary formats - can't have empty binary files

        if "empty" not in format_info["files"]:
            continue

        print(f"\nTesting empty file for format: {fmt}")
        empty_file = format_info["files"]["empty"]

        doc = parser.parse(empty_file)

        # Should still create valid document
        assert doc.document_id is not None
        assert doc.metadata is not None
        assert doc.processing_status == "completed"

        # Content should be empty
        assert doc.to_plain_text().strip() == ""
        assert len(doc.sections) == 0


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_flat_records_all_formats(
    parser,
    supported_format_files,
    parser_validation_suite,
):
    """Test flat record conversion for all supported formats."""
    validation = parser_validation_suite

    for fmt, format_info in supported_format_files.items():
        print(f"\nTesting flat records for format: {fmt}")

        # Use sample file for binary, simple file for text
        if format_info.get("is_binary", False):
            sample_files = [
                f
                for f in format_info["files"].values()
                if f.name.startswith("file_example")
            ]
            if not sample_files:
                continue
            file_path = sample_files[0]
        else:
            file_path = format_info["files"]["simple"]

        doc = parser.parse(file_path)
        records = doc.to_flat_records()

        # Validate record structure
        validation["validate_records"](records)

        # Check format-specific record expectations
        expectations = format_info["structure_expectations"]
        doc_records = [r for r in records if r["content_type"] == "document"]
        section_records = [r for r in records if r["content_type"] == "section"]
        para_records = [r for r in records if r["content_type"] == "paragraph"]

        assert len(doc_records) == 1
        if fmt in (".csv", ".xlsx"):
            table_records = [r for r in records if r.get("content_type") == "table"]
            assert (
                len(table_records) >= 1
            ), f"Expected at least one table record for {fmt}"
        else:
            if expectations["min_sections"] > 0:
                assert len(section_records) >= expectations["min_sections"]
            if expectations["min_paragraphs"] > 0:
                assert len(para_records) >= expectations["min_paragraphs"]


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_content_preservation_across_formats(parser, supported_format_files):
    """Test that content is properly preserved across different formats."""

    for fmt, format_info in supported_format_files.items():
        if format_info.get("is_binary", False):
            continue  # Binary file content depends on actual file content

        print(f"\nTesting content preservation for: {fmt}")

        # Test with different variants and their expected patterns
        for variant in ["simple", "multi_paragraph"]:
            if variant not in format_info["files"]:
                continue

            test_file = format_info["files"][variant]
            doc = parser.parse(test_file)
            full_text = doc.to_plain_text()

            # Should have substantial content
            assert (
                len(full_text.strip()) > 10
            ), f"Format {fmt} variant {variant} should extract substantial content"

            # Validate patterns based on the specific variant
            if variant == "simple" and format_info["validation_patterns"]:
                # For simple variant, check the validation patterns
                for pattern in format_info["validation_patterns"]:
                    # For CSV files, patterns are transformed (e.g., "Name,Age,City" becomes table format)
                    # So we check for individual words instead of exact comma-separated patterns
                    if fmt == ".csv":
                        # For CSV, check that individual column names exist (without commas)
                        words = pattern.replace(",", " ").split()
                        assert all(
                            word.lower() in full_text.lower() for word in words
                        ), f"CSV should contain all column names from pattern '{pattern}' in {fmt} {variant}"
                    else:
                        assert (
                            pattern.lower() in full_text.lower()
                        ), f"Pattern '{pattern}' not preserved in {fmt} {variant}"
            elif variant == "multi_paragraph":
                # For multi-paragraph, check for paragraph structure
                assert (
                    "paragraph" in full_text.lower()
                ), f"Multi-paragraph variant should contain 'paragraph' in {fmt}"
                assert (
                    len(full_text.split("\n\n")) >= 2
                ), f"Multi-paragraph should have multiple paragraphs in {fmt}"


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_metadata_consistency_across_formats(parser, supported_format_files):
    """Test that metadata is consistently set across all formats."""

    for fmt, format_info in supported_format_files.items():
        print(f"\nTesting metadata for format: {fmt}")

        # Use appropriate file for testing
        if format_info.get("is_binary", False):
            sample_files = [
                f
                for f in format_info["files"].values()
                if f.name.startswith("file_example")
            ]
            if not sample_files:
                continue
            file_path = sample_files[0]
        else:
            file_path = format_info["files"]["simple"]

        doc = parser.parse(file_path)

        # Check required metadata fields
        assert doc.metadata.mime_type.value == format_info["mime_type"]
        assert doc.metadata.parser_name == "DoclingParser"
        assert doc.metadata.parser_version is not None

        # Check file-specific metadata
        assert doc.metadata.file_name == file_path.name
        assert doc.metadata.file_size is not None
        assert doc.metadata.file_size >= 0

        # Check processing metadata
        assert doc.processing_status == "completed"
        assert doc.error_message is None


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_large_files_performance(
    parser,
    supported_format_files,
    performance_benchmarks,
):
    """Test performance with larger files where available."""
    benchmarks = performance_benchmarks

    for fmt, format_info in supported_format_files.items():
        if format_info.get("is_binary", False):
            continue  # Binary file size is fixed

        # Test with large variant if available
        if "large" not in format_info["files"]:
            continue

        print(f"\nTesting large file performance for: {fmt}")
        large_file = format_info["files"]["large"]

        # Get file size to determine benchmark
        file_size = large_file.stat().st_size
        if file_size < 1024:  # < 1KB
            max_time = benchmarks["max_parse_time"]["small_file"]
        elif file_size < 100 * 1024:  # < 100KB
            max_time = benchmarks["max_parse_time"]["medium_file"]
        else:
            max_time = benchmarks["max_parse_time"]["large_file"]

        # Time the parsing
        start_time = time.time()
        doc = parser.parse(large_file)
        parse_time = time.time() - start_time

        # Validate performance
        assert (
            parse_time <= max_time
        ), f"Parsing {fmt} took {parse_time:.2f}s, max allowed: {max_time}s"

        # Validate parsing succeeded
        assert doc.processing_status == "completed"
        assert len(doc.to_plain_text().strip()) > 100  # Should have substantial content


# ============================================================================
# CSV and Excel Format Tests
# ============================================================================


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_simple(parser, supported_format_files):
    """Test parsing simple CSV file with Docling's native support."""
    csv_files = supported_format_files[".csv"]["files"]
    csv_file = csv_files["simple"]
    doc = parser.parse(csv_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/csv"
    assert doc.metadata.file_name.endswith(".csv")
    assert doc.processing_status == "completed"

    # Check content is preserved - Docling should extract table data
    full_text = doc.to_plain_text()
    assert "John Doe" in full_text
    assert "Jane Smith" in full_text
    assert "New York" in full_text or "USA" in full_text

    # Should have structure (Docling extracts tables)
    assert len(doc.metadata.tables) >= 1, "CSV should produce at least one table"

    # Check basic statistics
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_simple")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_complex(parser, supported_format_files):
    """Test parsing complex CSV with quotes and special characters."""
    csv_files = supported_format_files[".csv"]["files"]
    csv_file = csv_files["complex"]
    doc = parser.parse(csv_file)

    # Check metadata
    assert doc.metadata.mime_type.value == "text/csv"
    assert doc.metadata.file_format.value == "csv"
    assert doc.processing_status == "completed"

    # Check content preservation with quotes
    full_text = doc.to_plain_text()
    assert "john.doe@company.com" in full_text.lower() or "Engineering" in full_text
    assert "Department" in full_text or "Salary" in full_text

    # Validate structure - CSV files produce tables
    assert len(doc.metadata.tables) >= 1
    # CSV files store data in tables, not paragraphs
    assert doc.metadata.total_characters > 0

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_complex")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_semicolon_delimiter(parser, tmp_path):
    """Test CSV with semicolon delimiter (Docling supports multiple delimiters)."""
    csv_file = tmp_path / "test_semicolon.csv"
    csv_content = """Name;Age;City
Alice;30;Paris
Bob;25;Berlin
Charlie;35;Madrid"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Docling should auto-detect semicolon delimiter
    assert doc.processing_status == "completed"

    full_text = doc.to_plain_text()
    assert "Alice" in full_text
    assert "Paris" in full_text
    assert "Berlin" in full_text

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file, sep=";")
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_semicolon_delimiter")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_pipe_delimiter(parser, tmp_path):
    """Test CSV with pipe delimiter (Docling supports |, ;, comma, tab)."""
    csv_file = tmp_path / "test_pipe.csv"
    csv_content = """Name|Department|Salary
John|Engineering|95000
Mary|Sales|75000
Steve|HR|65000"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Docling should auto-detect pipe delimiter
    assert doc.processing_status == "completed"

    full_text = doc.to_plain_text()
    assert "John" in full_text
    assert "Engineering" in full_text
    assert "95000" in full_text

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file, sep="|")
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_pipe_delimiter")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_with_unicode(parser, tmp_path):
    """Test CSV with Unicode characters."""
    csv_file = tmp_path / "test_unicode.csv"
    csv_content = """Name,City,Greeting
José,São Paulo,Olá
François,Montréal,Bonjour
李明,北京,你好"""
    csv_file.write_text(csv_content, encoding="utf-8")

    doc = parser.parse(str(csv_file))

    assert doc.processing_status == "completed"

    full_text = doc.to_plain_text()
    # Check for Unicode preservation
    assert "José" in full_text or "São Paulo" in full_text
    assert "François" in full_text or "Montréal" in full_text

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file, encoding="utf-8")
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_with_unicode")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_empty_cells(parser, tmp_path):
    """Test CSV with empty cells and sparse data."""
    csv_file = tmp_path / "test_empty.csv"
    csv_content = """Name,Age,City,Country
John,30,,USA
Jane,,London,
Bob,35,Sydney,Australia
,28,Toronto,Canada"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Should handle empty cells gracefully
    assert doc.processing_status == "completed"

    full_text = doc.to_plain_text()
    assert "John" in full_text
    assert "London" in full_text
    assert "Australia" in full_text

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_empty_cells")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_large_file_performance(parser, tmp_path):
    """Test parsing performance on larger CSV files."""
    import time

    csv_file = tmp_path / "large.csv"

    # Create a CSV with 1000 rows
    lines = ["ID,Name,Value,Category"]
    for i in range(1000):
        lines.append(f"{i},Item_{i},{i * 10.5},Category_{i % 10}")

    csv_file.write_text("\n".join(lines))

    start_time = time.time()
    doc = parser.parse(str(csv_file))
    parse_time = time.time() - start_time

    # Should complete in reasonable time (under 10 seconds)
    assert parse_time < 10.0, f"CSV parsing took {parse_time:.2f}s, too slow"
    assert doc.processing_status == "completed"

    # Verify content
    full_text = doc.to_plain_text()
    assert "Item_0" in full_text
    assert "Category" in full_text


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_to_schema_rows(parser, tmp_path):
    """Test that CSV documents can be converted to schema rows."""
    csv_file = tmp_path / "test_schema.csv"
    csv_content = """Employee,Department,Salary
Alice,Engineering,95000
Bob,Sales,75000"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Convert to schema rows
    rows = doc.to_schema_rows(document_index=0)

    # Should have document, section, paragraph, and sentence rows
    assert len(rows) > 0

    # Check row types
    row_types = [r["content_type"] for r in rows]
    assert "document" in row_types
    assert "table" in row_types

    # Check that content is present
    all_content = " ".join(r.get("content_text", "") for r in rows)
    assert "Alice" in all_content or "Engineering" in all_content

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_to_schema_rows")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_metadata_extraction(parser, tmp_path):
    """Test that CSV files have proper metadata extracted."""
    csv_file = tmp_path / "metadata_test.csv"
    csv_content = """Name,Value
Test,123
Data,456"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Check metadata fields
    assert doc.metadata.file_name == "metadata_test.csv"
    assert doc.metadata.mime_type.value == "text/csv"
    assert doc.metadata.file_format.value == "csv"
    assert doc.metadata.file_size > 0
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0
    assert doc.metadata.total_sections >= 0
    assert doc.metadata.parser_name == "DoclingParser"

    # Check timestamps are set
    assert doc.metadata.processed_at is not None
    assert doc.metadata.processing_time is not None
    assert doc.metadata.processing_time >= 0

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from CSV"

        df_expected = pd.read_csv(csv_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_csv_metadata_extraction")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_csv_document_structure(parser, tmp_path):
    """Test the hierarchical structure of parsed CSV documents."""
    csv_file = tmp_path / "structure_test.csv"
    csv_content = """Product,Price,Stock
Widget,19.99,100
Gadget,29.99,50
Gizmo,39.99,75"""
    csv_file.write_text(csv_content)

    doc = parser.parse(str(csv_file))

    # Check document structure
    assert doc.document_id is not None
    assert len(doc.document_id) > 0

    # CSVs produce tables (sections may be 0)
    assert len(doc.metadata.tables) > 0

    # For CSV files, validate document has content and table metadata
    assert doc.metadata.total_characters > 0
    assert len(doc.to_plain_text()) > 0

    # CSV files should have table metadata
    assert len(doc.metadata.tables) > 0, "CSV should produce table metadata"

    # Sections should have valid IDs
    for section in doc.sections:
        assert section.section_id is not None
        assert section.document_id == doc.document_id


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_workforce_data_xlsx(parser):
    """Test parsing workforce_data.xlsx with multiple sheets (Employees, Attendance, Salaries)."""
    from pathlib import Path

    # Use the actual XLSX file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    xlsx_file = sample_dir / "workforce_data.xlsx"

    if not xlsx_file.exists():
        pytest.skip("workforce_data.xlsx sample file not found")

    doc = parser.parse(xlsx_file)

    # Check metadata
    mt = doc.metadata.mime_type.value.lower()
    assert (
        "xlsx" in mt or "spreadsheet" in mt or "openxmlformats" in mt
    ), f"Expected XLSX file type, got: {doc.metadata.mime_type}"
    assert doc.metadata.file_name == "workforce_data.xlsx"
    assert doc.metadata.file_size > 0
    assert doc.processing_status == "completed"

    # Check content extraction - verify data from all three sheets
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "XLSX should contain extractable text"

    # ===== Sheet 1: Employees =====
    # Verify employee names (exact names from file)
    assert "Aria Patel" in full_text, "Should contain Aria Patel from Employees sheet"
    assert "Bilal Khan" in full_text, "Should contain Bilal Khan from Employees sheet"
    assert "Chen Li" in full_text, "Should contain Chen Li from Employees sheet"
    assert "Diego Reyes" in full_text, "Should contain Diego Reyes from Employees sheet"
    assert "Emma Novak" in full_text, "Should contain Emma Novak from Employees sheet"
    assert (
        "Farah Qureshi" in full_text
    ), "Should contain Farah Qureshi from Employees sheet"

    # Verify employee IDs from Employees sheet
    assert "301" in full_text, "Should contain EmployeeID 301"
    assert "302" in full_text, "Should contain EmployeeID 302"
    assert "303" in full_text, "Should contain EmployeeID 303"
    assert "304" in full_text, "Should contain EmployeeID 304"
    assert "305" in full_text, "Should contain EmployeeID 305"
    assert "306" in full_text, "Should contain EmployeeID 306"

    # Verify departments (exact departments from file)
    assert "Engineering" in full_text, "Should contain Engineering department"
    assert "Design" in full_text, "Should contain Design department"
    assert "Sales" in full_text, "Should contain Sales department"
    assert "Finance" in full_text, "Should contain Finance department"
    assert "HR" in full_text, "Should contain HR department"

    # ===== Sheet 2: Attendance =====
    # Verify attendance statuses (exact statuses from file)
    assert "Present" in full_text, "Should contain Present status from Attendance sheet"
    assert (
        "Absent" in full_text or "absent" in full_text.lower()
    ), "Should contain Absent status"
    assert (
        "Remote" in full_text or "remote" in full_text.lower()
    ), "Should contain Remote status"

    # Verify hours worked data
    assert "8" in full_text, "Should contain hours worked data (8 hours)"
    assert "0" in full_text, "Should contain 0 hours for absent employees"

    # ===== Sheet 3: Salaries =====
    # Verify salary data (exact values from file)
    assert (
        "98000" in full_text or "98,000" in full_text
    ), "Should contain salary 98000 (EmployeeID 301)"
    assert (
        "105000" in full_text or "105,000" in full_text
    ), "Should contain salary 105000 (EmployeeID 302)"
    assert (
        "86000" in full_text or "86,000" in full_text
    ), "Should contain salary 86000 (EmployeeID 303)"
    assert (
        "45000" in full_text or "45,000" in full_text
    ), "Should contain salary 45000 (EmployeeID 304)"
    assert (
        "92000" in full_text or "92,000" in full_text
    ), "Should contain salary 92000 (EmployeeID 305)"
    assert (
        "52000" in full_text or "52,000" in full_text
    ), "Should contain salary 52000 (EmployeeID 306)"

    # Verify bonus amounts (exact values from file)
    assert "10000" in full_text or "10,000" in full_text, "Should contain bonus 10000"
    assert "15000" in full_text or "15,000" in full_text, "Should contain bonus 15000"
    assert "8000" in full_text or "8,000" in full_text, "Should contain bonus 8000"
    assert "2000" in full_text or "2,000" in full_text, "Should contain bonus 2000"
    assert "12000" in full_text or "12,000" in full_text, "Should contain bonus 12000"
    assert "3000" in full_text or "3,000" in full_text, "Should contain bonus 3000"

    # Verify column headers across all sheets
    assert (
        "Employee" in full_text or "employee" in full_text.lower()
    ), "Should have Employee-related headers"
    assert "Department" in full_text, "Should have Department column"
    assert (
        "Salary" in full_text or "salary" in full_text.lower()
    ), "Should have Salary column"
    assert (
        "Attendance" in full_text
        or "attendance" in full_text.lower()
        or "Status" in full_text
    ), "Should reference Attendance/Status"
    assert (
        "Bonus" in full_text or "bonus" in full_text.lower()
    ), "Should have Bonus column"

    # Check structure - XLSX with multiple sheets should produce multiple tables
    assert len(doc.metadata.tables) >= 1, "XLSX should produce at least one table"

    # Check statistics
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0
    assert doc.metadata.total_sections >= 0

    # Validate table extraction - should have multiple tables for multiple sheets
    try:
        import pandas as pd

        # Check that tables were extracted
        assert (
            len(doc.metadata.tables) >= 3
        ), "Should extract at least 3 tables (one per sheet)"

        # Load each sheet with pandas and compare HTML
        sheet_names = ["Employees", "Attendance", "Salaries"]

        # For each sheet, load with pandas and verify EXACT HTML match
        for i, sheet_name in enumerate(sheet_names):
            df_expected = pd.read_excel(xlsx_file, sheet_name=sheet_name)
            expected_html = df_expected.to_html(index=False)
            normalized_expected = normalize_html_for_comparison(expected_html)

            # Get the corresponding extracted table HTML
            if i < len(doc.metadata.tables):
                docling_table_html = doc.metadata.tables[i].html
                assert (
                    docling_table_html is not None
                ), f"Table HTML for {sheet_name} should be extracted"
                normalized_docling = normalize_html_for_comparison(docling_table_html)

                # Use LLM judge for semantic HTML equivalence
                is_equivalent, explanation = await llm_judge_html_equivalence(
                    normalized_expected,
                    normalized_docling,
                )
                assert is_equivalent, (
                    f"HTML tables are not semantically equivalent for sheet '{sheet_name}'!\n"
                    f"Reason: {explanation}\n\n"
                    f"Expected HTML:\n{normalized_expected}\n\n"
                    f"Parsed HTML:\n{normalized_docling}"
                )

    except ImportError:
        # If pandas is not available, we still pass the test since core assertions passed
        print("Skipping pandas verification for workforce_data.xlsx")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_retail_data_xlsx(parser):
    """Test parsing retail_data.xlsx with multiple sheets (Stores, Sales, Inventory, Returns)."""
    from pathlib import Path

    # Use the actual XLSX file from sample directory
    sample_dir = Path(__file__).parent.parent / "sample"
    xlsx_file = sample_dir / "retail_data.xlsx"

    if not xlsx_file.exists():
        pytest.skip("retail_data.xlsx sample file not found")

    doc = parser.parse(xlsx_file)

    # Check metadata
    mt = doc.metadata.mime_type.value.lower()
    assert (
        "xlsx" in mt or "spreadsheet" in mt or "openxmlformats" in mt
    ), f"Expected XLSX file type, got: {doc.metadata.mime_type}"
    assert doc.metadata.file_name == "retail_data.xlsx"
    assert doc.metadata.file_size > 0
    assert doc.processing_status == "completed"

    # Check content extraction - verify data from all four sheets
    full_text = doc.to_plain_text()
    assert len(full_text.strip()) > 0, "XLSX should contain extractable text"

    # ===== Sheet 1: Stores =====
    # Verify store names (exact names from file)
    assert "Gulshan" in full_text, "Should contain Gulshan store from Stores sheet"
    assert "DHA" in full_text, "Should contain DHA store"
    assert "Blue Area" in full_text, "Should contain Blue Area store"
    assert "Saddar" in full_text, "Should contain Saddar store"

    # Verify cities (exact cities from file)
    assert "Karachi" in full_text, "Should contain Karachi city"
    assert "Lahore" in full_text, "Should contain Lahore city"
    assert "Islamabad" in full_text, "Should contain Islamabad city"
    assert "Rawalpindi" in full_text, "Should contain Rawalpindi city"

    # Verify store IDs (exact IDs from file)
    assert "10" in full_text, "Should contain StoreID 10 (Gulshan)"
    assert "11" in full_text, "Should contain StoreID 11 (DHA)"
    assert "12" in full_text, "Should contain StoreID 12 (Blue Area)"
    assert "13" in full_text, "Should contain StoreID 13 (Saddar)"

    # ===== Sheet 2: Sales =====
    # Verify sale IDs (exact IDs from file)
    assert "5001" in full_text, "Should contain SaleID 5001 from Sales sheet"
    assert "5002" in full_text, "Should contain SaleID 5002"
    assert "5003" in full_text, "Should contain SaleID 5003"
    assert "5004" in full_text, "Should contain SaleID 5004"
    assert "5005" in full_text, "Should contain SaleID 5005"
    assert "5006" in full_text, "Should contain SaleID 5006"

    # Verify SKUs (exact SKUs from file)
    assert (
        "LTP-15" in full_text or "ltp-15" in full_text.lower()
    ), "Should contain SKU LTP-15 (Laptop)"
    assert (
        "MOU-01" in full_text or "mou-01" in full_text.lower()
    ), "Should contain SKU MOU-01 (Mouse)"
    assert (
        "KBD-02" in full_text or "kbd-02" in full_text.lower()
    ), "Should contain SKU KBD-02 (Keyboard)"
    assert (
        "MON-27" in full_text or "mon-27" in full_text.lower()
    ), "Should contain SKU MON-27 (Monitor)"
    assert (
        "PRN-10" in full_text or "prn-10" in full_text.lower()
    ), "Should contain SKU PRN-10 (Printer)"

    # Verify amounts in PKR (exact amounts from file)
    assert (
        "420000" in full_text or "420,000" in full_text
    ), "Should contain amount 420000 PKR"
    assert "7500" in full_text or "7,500" in full_text, "Should contain amount 7500 PKR"
    assert (
        "12000" in full_text or "12,000" in full_text
    ), "Should contain amount 12000 PKR"
    assert (
        "85000" in full_text or "85,000" in full_text
    ), "Should contain amount 85000 PKR"
    assert (
        "45000" in full_text or "45,000" in full_text
    ), "Should contain amount 45000 PKR"
    assert "6000" in full_text or "6,000" in full_text, "Should contain amount 6000 PKR"

    # ===== Sheet 3: Inventory =====
    # Verify item names (exact names from file)
    assert (
        "Laptop" in full_text or "laptop" in full_text.lower()
    ), 'Should contain Laptop 15" from Inventory'
    assert (
        "Mouse" in full_text or "mouse" in full_text.lower()
    ), "Should contain Wireless Mouse"
    assert (
        "Keyboard" in full_text or "keyboard" in full_text.lower()
    ), "Should contain Mechanical Keyboard"
    assert (
        "Monitor" in full_text or "monitor" in full_text.lower()
    ), 'Should contain Monitor 27"'
    assert (
        "Printer" in full_text or "printer" in full_text.lower()
    ), "Should contain Laser Printer"

    # Verify quantity data (exact quantities from file)
    assert "7" in full_text, "Should contain quantity 7 (LTP-15 at Store 10)"
    assert "120" in full_text, "Should contain quantity 120 (MOU-01 at Store 11)"
    assert "45" in full_text, "Should contain quantity 45 (KBD-02 at Store 12)"
    assert "12" in full_text, "Should contain quantity 12 (MON-27 at Store 13)"
    assert "6" in full_text, "Should contain quantity 6 (PRN-10 at Store 10)"

    # ===== Sheet 4: Returns =====
    # Verify return data (exact IDs from file)
    assert "9001" in full_text, "Should contain ReturnID 9001 from Returns sheet"
    assert "9002" in full_text, "Should contain ReturnID 9002"

    # Verify return reasons (exact reasons from file)
    assert (
        "Defective" in full_text or "defective" in full_text.lower()
    ), "Should contain return reason 'Defective mouse'"
    assert (
        "Damaged" in full_text or "damaged" in full_text.lower()
    ), "Should contain return reason 'Damaged packaging'"

    # Verify refund amounts (exact amounts from file)
    assert "1500" in full_text, "Should contain refund 1500 PKR (ReturnID 9001)"
    assert "5000" in full_text, "Should contain refund 5000 PKR (ReturnID 9002)"

    # Verify column headers across all sheets
    assert (
        "Store" in full_text or "store" in full_text.lower()
    ), "Should have Store-related headers"
    assert "Sale" in full_text or "sale" in full_text.lower(), "Should have Sale column"
    assert (
        "Inventory" in full_text or "inventory" in full_text.lower()
    ), "Should reference Inventory"
    assert (
        "Return" in full_text or "return" in full_text.lower()
    ), "Should have Return column"
    assert "SKU" in full_text or "sku" in full_text.lower(), "Should have SKU column"

    # Check structure - XLSX with multiple sheets should produce multiple tables
    assert len(doc.metadata.tables) >= 1, "XLSX should produce at least one table"

    # Check statistics
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_words > 0
    assert doc.metadata.total_sections >= 0

    # Validate table extraction - should have multiple tables for multiple sheets
    try:
        import pandas as pd

        # Check that tables were extracted
        assert (
            len(doc.metadata.tables) >= 4
        ), "Should extract at least 4 tables (one per sheet)"

        # Load each sheet with pandas and compare HTML
        sheet_names = ["Stores", "Sales", "Inventory", "Returns"]

        # For each sheet, load with pandas and verify EXACT HTML match
        for i, sheet_name in enumerate(sheet_names):
            df_expected = pd.read_excel(xlsx_file, sheet_name=sheet_name)
            expected_html = df_expected.to_html(index=False)
            normalized_expected = normalize_html_for_comparison(expected_html)

            # Get the corresponding extracted table HTML
            if i < len(doc.metadata.tables):
                docling_table_html = doc.metadata.tables[i].html
                assert (
                    docling_table_html is not None
                ), f"Table HTML for {sheet_name} should be extracted"
                normalized_docling = normalize_html_for_comparison(docling_table_html)

                # Use LLM judge for semantic HTML equivalence
                is_equivalent, explanation = await llm_judge_html_equivalence(
                    normalized_expected,
                    normalized_docling,
                )
                assert is_equivalent, (
                    f"HTML tables are not semantically equivalent for sheet '{sheet_name}'!\n"
                    f"Reason: {explanation}\n\n"
                    f"Expected HTML:\n{normalized_expected}\n\n"
                    f"Parsed HTML:\n{normalized_docling}"
                )

    except ImportError:
        # If pandas is not available, we still pass the test since core assertions passed
        print("Skipping pandas verification for retail_data.xlsx")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_xlsx_multiple_sheets(parser, tmp_path):
    """Test XLSX with multiple sheets."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    xlsx_file = tmp_path / "multi_sheet.xlsx"

    wb = Workbook()

    # Sheet 1: Sales
    ws1 = wb.active
    ws1.title = "Q1 Sales"
    ws1["A1"] = "Product"
    ws1["B1"] = "Revenue"
    ws1["A2"] = "Widget"
    ws1["B2"] = 50000
    ws1["A3"] = "Gadget"
    ws1["B3"] = 75000

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("Q1 Expenses")
    ws2["A1"] = "Category"
    ws2["B1"] = "Amount"
    ws2["A2"] = "Salaries"
    ws2["B2"] = 30000
    ws2["A3"] = "Marketing"
    ws2["B3"] = 15000

    wb.save(str(xlsx_file))

    # Parse
    doc = parser.parse(str(xlsx_file))

    assert doc.processing_status == "completed"

    # Should extract content from both sheets
    full_text = doc.to_plain_text()
    assert "Widget" in full_text or "Gadget" in full_text
    assert "Salaries" in full_text or "Marketing" in full_text

    # May have multiple sections for multiple sheets
    assert len(doc.sections) >= 0

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert (
            len(doc.metadata.tables) >= 2
        ), "Should extract at least 2 tables (one per sheet)"

        # Check first sheet (Q1 Sales) - Use LLM judge for equivalence
        df1 = pd.read_excel(xlsx_file, sheet_name="Q1 Sales")
        expected_html1 = df1.to_html(index=False)
        normalized_expected1 = normalize_html_for_comparison(expected_html1)

        if len(doc.metadata.tables) > 0:
            docling_html1 = doc.metadata.tables[0].html
            assert docling_html1 is not None
            normalized_docling1 = normalize_html_for_comparison(docling_html1)
            is_equivalent, explanation = await llm_judge_html_equivalence(
                normalized_expected1,
                normalized_docling1,
            )
            assert is_equivalent, (
                f"HTML tables are not semantically equivalent for sheet 'Q1 Sales'!\n"
                f"Reason: {explanation}\n\n"
                f"Expected HTML:\n{normalized_expected1}\n\n"
                f"Parsed HTML:\n{normalized_docling1}"
            )

        # Check second sheet (Q1 Expenses) - Use LLM judge for equivalence
        df2 = pd.read_excel(xlsx_file, sheet_name="Q1 Expenses")
        expected_html2 = df2.to_html(index=False)
        normalized_expected2 = normalize_html_for_comparison(expected_html2)

        if len(doc.metadata.tables) > 1:
            docling_html2 = doc.metadata.tables[1].html
            assert docling_html2 is not None
            normalized_docling2 = normalize_html_for_comparison(docling_html2)
            is_equivalent, explanation = await llm_judge_html_equivalence(
                normalized_expected2,
                normalized_docling2,
            )
            assert is_equivalent, (
                f"HTML tables are not semantically equivalent for sheet 'Q1 Expenses'!\n"
                f"Reason: {explanation}\n\n"
                f"Expected HTML:\n{normalized_expected2}\n\n"
                f"Parsed HTML:\n{normalized_docling2}"
            )

    except ImportError:
        print("Skipping pandas verification for test_xlsx_multiple_sheets")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_xlsx_with_formulas(parser, tmp_path):
    """Test XLSX with formulas."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    xlsx_file = tmp_path / "formulas.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"

    ws["A1"] = "Value1"
    ws["B1"] = "Value2"
    ws["C1"] = "Sum"

    ws["A2"] = 100
    ws["B2"] = 200
    ws["C2"] = "=A2+B2"  # Formula

    ws["A3"] = 50
    ws["B3"] = 75
    ws["C3"] = "=A3+B3"  # Formula

    wb.save(str(xlsx_file))

    # Parse
    doc = parser.parse(str(xlsx_file))

    assert doc.processing_status == "completed"

    # Check that at least the input values are present
    full_text = doc.to_plain_text()
    assert "100" in full_text
    assert "200" in full_text
    # Docling may or may not evaluate formulas, but should extract something
    assert len(full_text) > 50

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from XLSX"

        # Load with pandas (formulas won't be evaluated in pandas either without engine)
        df_expected = pd.read_excel(xlsx_file, sheet_name="Calculations")
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_xlsx_with_formulas")


@pytest.mark.asyncio
@pytest.mark.unit
@_handle_project
async def test_xlsx_metadata_extraction(parser, tmp_path):
    """Test metadata extraction from XLSX files."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    xlsx_file = tmp_path / "metadata_test.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Test"
    ws["B2"] = 123

    wb.save(str(xlsx_file))

    doc = parser.parse(str(xlsx_file))

    # Check all metadata fields
    assert doc.metadata.file_name == "metadata_test.xlsx"
    mt = doc.metadata.mime_type.value.lower()
    assert "xlsx" in mt or "spreadsheet" in mt
    assert doc.metadata.file_size > 0
    assert doc.metadata.total_characters > 0
    assert doc.metadata.total_sections >= 0
    assert doc.metadata.parser_name == "DoclingParser"
    assert doc.metadata.processing_time is not None
    assert doc.metadata.processing_time >= 0

    # Validate table extraction with pandas HTML comparison
    try:
        import pandas as pd

        assert len(doc.metadata.tables) > 0, "Should extract table metadata from XLSX"

        df_expected = pd.read_excel(xlsx_file)
        expected_html = df_expected.to_html(index=False)

        docling_table_html = doc.metadata.tables[0].html
        assert docling_table_html is not None, "Table HTML should be extracted"

        normalized_expected = normalize_html_for_comparison(expected_html)
        normalized_docling = normalize_html_for_comparison(docling_table_html)

        # Use LLM judge for semantic HTML equivalence
        is_equivalent, explanation = await llm_judge_html_equivalence(
            normalized_expected,
            normalized_docling,
        )
        assert is_equivalent, (
            f"HTML tables are not semantically equivalent!\n"
            f"Reason: {explanation}\n\n"
            f"Expected HTML:\n{normalized_expected}\n\n"
            f"Parsed HTML:\n{normalized_docling}"
        )

    except ImportError:
        print("Skipping pandas verification for test_xlsx_metadata_extraction")
