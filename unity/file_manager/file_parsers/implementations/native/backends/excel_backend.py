from __future__ import annotations

import logging
import time
from collections.abc import Iterator
from pathlib import Path
from typing import Optional, Sequence

from openpyxl import load_workbook

from unity.file_manager.file_parsers.implementations.native.spreadsheet_support import (
    finalize_spreadsheet_result,
    normalize_tabular_value,
    should_inline_tabular_rows,
)
from unity.file_manager.file_parsers.settings import FILE_PARSER_SETTINGS
from unity.file_manager.file_parsers.types.backend import BaseFileParserBackend
from unity.file_manager.file_parsers.types.contracts import (
    FileParseRequest,
    FileParseResult,
    FileParseTrace,
    StepStatus,
)
from unity.file_manager.file_parsers.types.formats import FileFormat
from unity.file_manager.file_parsers.types.json_types import JsonObject
from unity.file_manager.file_parsers.types.table import ExtractedTable
from unity.file_manager.file_parsers.utils.tracing import traced_step

logger = logging.getLogger(__name__)


class NativeExcelBackend(BaseFileParserBackend):
    """Streaming XLSX backend backed by openpyxl read-only worksheets."""

    name = "native_excel_backend"
    supported_formats: Sequence[FileFormat] = (FileFormat.XLSX,)

    def can_handle(self, fmt: Optional[FileFormat]) -> bool:
        return fmt in self.supported_formats

    @staticmethod
    def iter_rows(
        path: str | Path,
        *,
        sheet_name: str | None = None,
        has_header: bool = True,
        columns: Sequence[str] | None = None,
    ) -> "Iterator[dict[str, object]]":
        """Stream rows from an XLSX sheet via openpyxl read-only mode.

        This is the single row-iteration entry point for Excel files.
        Callers that hold an ``XlsxSheetHandle`` should unpack it into
        keyword arguments and call this directly.
        """
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        try:
            ws = workbook[sheet_name] if sheet_name else workbook.active
            rows_iter = ws.iter_rows(values_only=True)

            col_names = list(columns or [])
            if has_header:
                header = next(rows_iter, None)
                if header is None:
                    return
                if not col_names:
                    col_names = _coerce_header_row(
                        tuple(normalize_tabular_value(v) for v in header),
                    )
            elif not col_names:
                first = next(rows_iter, None)
                if first is None:
                    return
                col_names = [f"column_{i}" for i in range(1, len(first) + 1)]
                yield _row_from_values(
                    col_names,
                    [normalize_tabular_value(v) for v in first],
                )

            for row in rows_iter:
                values = [normalize_tabular_value(v) for v in row]
                if _is_blank_row(values):
                    continue
                yield _row_from_values(col_names, values)
        finally:
            workbook.close()

    def parse(self, ctx: FileParseRequest, /) -> FileParseResult:
        started = time.perf_counter()
        path = Path(ctx.source_local_path).expanduser().resolve()

        trace = FileParseTrace(
            logical_path=str(ctx.logical_path),
            backend=self.name,
            file_format=ctx.file_format,
            mime_type=ctx.mime_type,
            status=StepStatus.SUCCESS,
            source_local_path=str(path),
            parsed_local_path=str(path),
        )

        if not path.exists() or not path.is_file():
            trace.status = StepStatus.FAILED
            trace.duration_ms = (time.perf_counter() - started) * 1000.0
            return FileParseResult(
                logical_path=str(ctx.logical_path),
                status="error",
                error=f"File not found: {path}",
                file_format=ctx.file_format,
                mime_type=ctx.mime_type,
                trace=trace,
            )

        try:
            with traced_step(trace, name="stream_workbook_sheets") as step:
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    tables: list[ExtractedTable] = []
                    sheet_names: list[str] = []
                    for sheet_index, worksheet in enumerate(
                        workbook.worksheets,
                        start=1,
                    ):
                        table = _parse_worksheet(
                            worksheet=worksheet,
                            sheet_index=sheet_index,
                        )
                        if table is None:
                            continue
                        tables.append(table)
                        sheet_names.append(
                            str(table.sheet_name or f"Sheet {sheet_index}"),
                        )
                    step.counters["worksheets"] = len(workbook.worksheets)
                    step.counters["tables"] = len(tables)
                finally:
                    workbook.close()

            result = finalize_spreadsheet_result(
                logical_path=str(ctx.logical_path),
                file_format=ctx.file_format,
                mime_type=ctx.mime_type,
                trace=trace,
                settings=FILE_PARSER_SETTINGS,
                tables=tables,
                sheet_names=sheet_names,
            )
            trace.duration_ms = (time.perf_counter() - started) * 1000.0
            return result

        except Exception as e:
            logger.exception("Native XLSX parse failed: %s", e)
            trace.status = StepStatus.FAILED
            trace.warnings.append(str(e))
            trace.duration_ms = (time.perf_counter() - started) * 1000.0
            return FileParseResult(
                logical_path=str(ctx.logical_path),
                status="error",
                error=str(e),
                file_format=ctx.file_format,
                mime_type=ctx.mime_type,
                trace=trace,
            )


def _parse_worksheet(*, worksheet, sheet_index: int) -> ExtractedTable | None:
    rows_iter = worksheet.iter_rows(values_only=True)
    header = _first_non_empty_row(rows_iter)
    if header is None:
        return None

    columns = _coerce_header_row(header)
    row_count = 0
    sample_rows: list[JsonObject] = []
    inline_rows: list[JsonObject] = []

    for row in rows_iter:
        values = [normalize_tabular_value(value) for value in row]
        if _is_blank_row(values):
            continue
        record = _row_from_values(columns, values)
        row_count += 1
        if len(sample_rows) < int(FILE_PARSER_SETTINGS.TABULAR_SAMPLE_ROWS):
            sample_rows.append(record)
        if should_inline_tabular_rows(
            row_count=row_count,
            settings=FILE_PARSER_SETTINGS,
        ):
            inline_rows.append(record)

    if not should_inline_tabular_rows(
        row_count=row_count,
        settings=FILE_PARSER_SETTINGS,
    ):
        inline_rows = []

    sheet_name = (
        str(worksheet.title or f"Sheet {sheet_index}").strip() or f"Sheet {sheet_index}"
    )
    return ExtractedTable(
        table_id=f"table:{sheet_index}",
        label=sheet_name,
        sheet_name=sheet_name,
        columns=columns,
        rows=inline_rows,
        sample_rows=sample_rows,
        num_rows=row_count,
        num_cols=len(columns),
    )


def _first_non_empty_row(rows_iter) -> tuple[object, ...] | None:
    for row in rows_iter:
        values = [normalize_tabular_value(value) for value in row]
        if _is_blank_row(values):
            continue
        return tuple(values)
    return None


def _coerce_header_row(header: tuple[object, ...]) -> list[str]:
    columns: list[str] = []
    for index, value in enumerate(header, start=1):
        text = str(value).strip() if value is not None else ""
        columns.append(text or f"column_{index}")
    return columns


def _row_from_values(columns: list[str], values: list[object]) -> JsonObject:
    row: JsonObject = {}
    for index, column in enumerate(columns):
        row[str(column)] = values[index] if index < len(values) else None
    return row


def _is_blank_row(values: list[object]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True
