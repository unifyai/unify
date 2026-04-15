from __future__ import annotations

import csv
import json
from collections.abc import Iterator
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

from unity.file_manager.file_parsers.types.json_types import JsonObject

from .types import (
    CsvFileHandle,
    InlineRowsHandle,
    ObjectStoreArtifactHandle,
    TableInputHandle,
    XlsxSheetHandle,
)


def iter_table_input_rows(handle: TableInputHandle) -> Iterator[JsonObject]:
    """Yield table rows from any supported transport handle."""

    if isinstance(handle, InlineRowsHandle):
        yield from (dict(row) for row in handle.rows)
        return

    if isinstance(handle, CsvFileHandle):
        yield from _iter_csv_rows(handle)
        return

    if isinstance(handle, XlsxSheetHandle):
        yield from _iter_xlsx_rows(handle)
        return

    if isinstance(handle, ObjectStoreArtifactHandle):
        yield from _iter_object_store_rows(handle)
        return

    raise TypeError(f"Unsupported table input handle: {type(handle)!r}")


def iter_table_input_row_batches(
    handle: TableInputHandle,
    batch_size: int,
) -> Iterator[list[JsonObject]]:
    """Yield bounded row batches from a table input handle."""

    size = max(int(batch_size or 0), 1)
    batch: list[JsonObject] = []
    for row in iter_table_input_rows(handle):
        batch.append(row)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def _iter_csv_rows(handle: CsvFileHandle) -> Iterator[JsonObject]:
    path = _resolve_local_path(
        source_local_path=handle.source_local_path,
        storage_uri=handle.storage_uri,
    )
    with path.open("r", encoding=handle.encoding, newline="") as fh:
        reader = csv.reader(
            fh,
            delimiter=handle.delimiter,
            quotechar=handle.quotechar,
        )

        columns = list(handle.columns or [])
        if handle.has_header:
            first_row = next(reader, None)
            if first_row is None:
                return
            if not columns:
                columns = [str(cell) for cell in first_row]
        elif not columns:
            first_row = next(reader, None)
            if first_row is None:
                return
            columns = [f"column_{idx + 1}" for idx in range(len(first_row))]
            yield _row_from_pairs(columns, first_row)

        for row in reader:
            yield _row_from_pairs(columns, row)


def _iter_xlsx_rows(handle: XlsxSheetHandle) -> Iterator[JsonObject]:
    path = _resolve_local_path(
        source_local_path=handle.source_local_path,
        storage_uri=handle.storage_uri,
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[handle.sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        columns = list(handle.columns or [])

        if handle.has_header:
            header = next(rows, None)
            if header is None:
                return
            if not columns:
                columns = [
                    _normalize_excel_cell(v) if v is not None else "" for v in header
                ]
                columns = [str(col) for col in columns]
        elif not columns:
            first_row = next(rows, None)
            if first_row is None:
                return
            columns = [f"column_{idx + 1}" for idx in range(len(first_row))]
            yield _row_from_pairs(
                columns,
                [_normalize_excel_cell(v) for v in first_row],
            )

        for row in rows:
            yield _row_from_pairs(columns, [_normalize_excel_cell(v) for v in row])
    finally:
        workbook.close()


def _iter_object_store_rows(handle: ObjectStoreArtifactHandle) -> Iterator[JsonObject]:
    if handle.artifact_format != "jsonl":
        raise NotImplementedError(
            f"Artifact streaming is not implemented for {handle.artifact_format!r}",
        )

    path = _resolve_local_path(source_local_path="", storage_uri=handle.storage_uri)
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            text = line.strip()
            if not text:
                continue
            loaded = json.loads(text)
            if isinstance(loaded, dict):
                yield {str(key): value for key, value in loaded.items()}


def _row_from_pairs(columns: list[str], values: list[object]) -> JsonObject:
    row: JsonObject = {}
    if not columns and values:
        columns = [f"column_{idx + 1}" for idx in range(len(values))]
    for idx, column in enumerate(columns):
        row[str(column)] = values[idx] if idx < len(values) else None
    return row


def _normalize_excel_cell(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _resolve_local_path(*, source_local_path: str, storage_uri: str) -> Path:
    if source_local_path:
        return Path(source_local_path).expanduser().resolve()
    parsed = urlparse(storage_uri)
    if parsed.scheme == "file":
        return Path(unquote(parsed.path)).expanduser().resolve()
    raise ValueError(f"Cannot resolve local path from storage URI: {storage_uri}")
