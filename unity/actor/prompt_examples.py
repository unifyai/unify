"""
Centralized example library for Actor prompt builders.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import List


@dataclass
class Example:
    """Container for a single example with metadata."""

    title: str
    code: str
    description: str
    environment_tags: List[str]  # e.g., ["computer"], ["primitives"], ["mixed"]


# ---------------------------------------------------------------------------
# 2. Core Patterns (Environment-Agnostic)
# ---------------------------------------------------------------------------


def get_confidence_based_stubbing_example() -> str:
    """Example: stub uncertain steps with clear TODO markers."""

    return """
# Example: Confidence-based stubbing
async def main_plan():
    # High confidence: direct implementation (tool-domain agnostic)
    records = await fetch_records()

    # Low confidence: stub with clear marker
    def send_email_to_contacts(_records: list) -> str:
        # TODO (LOW confidence): decide email service + template + recipients policy.
        raise NotImplementedError("Stub: email sending not yet implemented")

    # Continue with high-confidence steps
    result = send_email_to_contacts(records)
    return result
"""


def get_structured_output_example() -> str:
    """Example: use Pydantic for structured extraction."""

    return """
# Example: Structured output with Pydantic
from pydantic import BaseModel
from typing import Optional

class ContactInfo(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None

async def extract_structured(query: str) -> ContactInfo:
    # In practice: pass response_format=ContactInfo to a tool call and return the model.
    return ContactInfo(name="Alice Smith", email="alice.smith@example.com", phone="1112223333")
"""


def get_error_handling_example() -> str:
    """Example: graceful error handling with fallbacks."""

    return """
# Example: Error handling with fallbacks
async def find_with_fallback(query: str) -> str:
    try:
        # Primary: strict query
        out = await strict_lookup(query)
        if "not found" in str(out).lower():
            raise ValueError("Primary lookup failed")
        return str(out)
    except Exception as e:
        # Fallback: broader query
        out2 = await broad_lookup(query)
        return f"Fallback: {out2}"
"""


def get_handle_steering_example() -> str:
    """Example: steering in-flight handles."""

    return """
# Example: Handle steering (pause/resume/interject)
async def run_with_steering() -> str:
    # Start a long-running tool loop (returns steerable handle)
    handle = await start_long_running_work()

    # Pause execution to review progress
    await handle.pause()

    # Query current status
    status_handle = await handle.ask("What is the current progress?")
    status = await status_handle.result()

    # Inject guidance based on status
    await handle.interject("Prioritize data validation before proceeding.")

    # Resume execution
    await handle.resume()

    # Wait for completion
    result = await handle.result()
    return result
"""


def get_handle_mode_selection_example() -> str:
    """Example: choosing execute_function vs execute_code."""

    return """
# Example: execute_function vs execute_code decision
#
# CORRECT — single primitive call → use execute_function (JSON tool call):
#   execute_function(function_name="primitives.contacts.ask", call_kwargs={"text": "Find contacts in Berlin"})
#   execute_function(function_name="primitives.tasks.execute", call_kwargs={"task_id": 123})
#   execute_function(function_name="primitives.knowledge.update", call_kwargs={"text": "Store: Office hours are 9-5 PT."})
#
# WRONG — wrapping a single primitive in execute_code just to add notify():
#   execute_code(code='''
#       notify({"type": "progress", "message": "Looking up contacts..."})
#       handle = await primitives.contacts.ask(text="Find contacts in Berlin")
#       result = await handle.result()
#       notify({"type": "step_complete", ...})
#       print(result)
#   ''')
#   ^^^ This consumes the handle inside the code block. The outer loop
#       loses steering and query access. Use execute_function instead.
#
# CORRECT — genuine multi-step composition requires execute_code:
async def cross_reference_contacts_and_transcripts(city: str) -> str:
    notify({
        "type": "progress",
        "message": f"Step 1: Fetching contacts in {city}.",
        "step": 1,
        "total": 2
    })
    contacts_handle = await primitives.contacts.ask(f"List contacts in {city}.")
    contacts = await contacts_handle.result()

    notify({
        "type": "progress",
        "message": "Step 2: Summarizing recent interactions for those contacts.",
        "step": 2,
        "total": 2
    })
    transcript_handle = await primitives.transcripts.ask(
        f"Summarize recent interactions for contacts in {city}."
    )
    summary = await transcript_handle.result()

    return f"Contacts: {contacts}\\nInteractions: {summary}"
"""


def get_clarification_example() -> str:
    """Example: handling clarification requests from tools."""

    return """
# Example: Clarification requests
async def tool_call_that_may_clarify() -> str:
    # Some tool calls may request clarification mid-flight; the Actor/pane can:
    # - answer immediately if low-risk and known
    # - pause and ask the user for a disambiguating choice
    handle = await start_mutation_that_may_clarify()
    return await handle.result()
"""


def get_notify_web_search_example() -> str:
    """Example: single web search vs multi-step research."""

    return """
# Example: Single web search → execute_function (NOT execute_code)
#
# CORRECT (JSON tool call):
#   execute_function(function_name="primitives.web.ask",
#                    call_kwargs={"text": "What is the weather in Berlin today?"})
#
# WRONG — wrapping a single web.ask in execute_code:
#   execute_code(code='''
#       notify(...)
#       handle = await primitives.web.ask("What is the weather in Berlin today?")
#       result = await handle.result()
#   ''')
#
# Multi-step research (genuinely needs execute_code):
async def gather_and_verify_role_openings(query: str) -> str:
    notify({
        "type": "progress",
        "message": "Searching public sources for relevant role listings.",
        "step": 1,
        "total": 2
    })

    initial_handle = await primitives.web.ask(
        f"Find current openings related to: {query}"
    )
    initial_results = await initial_handle.result()

    notify({
        "type": "progress",
        "message": "Validating listings against official company pages.",
        "step": 2,
        "total": 2,
    })

    verified_handle = await primitives.web.ask(
        f"Cross-check these listings against official careers pages: {initial_results}"
    )
    verified_results = await verified_handle.result()

    return verified_results
"""


def get_notify_multistep_workflow_example() -> str:
    """Example: using notify() between meaningful workflow milestones."""

    return """
# Example: Multi-step workflow (genuinely compositional → execute_code is correct)
# NOTE: If this were a single primitive call, use execute_function instead.
async def build_contact_insights(city: str) -> str:
    notify({
        "type": "progress",
        "message": f"Fetching contacts for {city}.",
        "step": 1,
        "total": 3
    })
    contacts_handle = await primitives.contacts.ask(
        f"List contacts in {city} with role and company."
    )
    contacts = await contacts_handle.result()

    notify({
        "type": "progress",
        "message": "Summarizing recent interactions for matching contacts.",
        "step": 2,
        "total": 3
    })
    transcript_handle = await primitives.transcripts.ask(
        f"Summarize recent interactions for contacts in {city}."
    )
    interaction_summary = await transcript_handle.result()

    notify({
        "type": "progress",
        "message": "Persisting synthesized insights to the knowledge store.",
        "step": 3,
        "total": 3
    })
    save_handle = await primitives.knowledge.update(
        f"Store structured contact insights for {city}: contacts={contacts}, summary={interaction_summary}"
    )
    save_result = await save_handle.result()

    return str(save_result)
"""


def get_notify_long_running_example() -> str:
    """Example: periodic notify() usage for long-running operations."""

    return """
# Example: Long-running operation with concrete periodic notifications
async def process_large_collection(records: list[dict]) -> dict:
    batch_size = 250
    total_batches = max(1, (len(records) + batch_size - 1) // batch_size)

    processed = 0
    rejected = 0

    for index in range(total_batches):
        start = index * batch_size
        end = min(len(records), start + batch_size)
        batch = records[start:end]

        notify({
            "type": "progress",
            "message": f"Processing batch {index + 1}/{total_batches} ({processed} processed, {rejected} rejected so far).",
            "step": index + 1,
            "total": total_batches,
        })

        for item in batch:
            if item.get("is_valid"):
                processed += 1
            else:
                rejected += 1

    # GOOD: specific, measurable updates
    # BAD: generic filler messages with no new signal:
    # notify({"type": "progress", "message": "Working on it..."})
    # notify({"type": "progress", "message": "Still processing..."})
    # BAD: low-level internal diagnostics instead of user-facing progress:
    # notify({"type": "progress", "message": "tool_call_id=abc123, state_mode=stateful, parser=ProductList"})
    return {"processed": processed, "rejected": rejected, "total": len(records)}
"""


def get_library_function_reuse_example() -> str:
    """Example: recognizing when to call an existing library function directly.

    Shows the pattern of checking library skills and calling them when they match the goal.
    """

    return """
# Example: Reusing an existing library function
async def main_plan():
    '''User goal: Schedule a weekly team meeting every Monday at 9am.

    ANALYSIS: The library has a skill 'schedule_recurring_task' that handles
    recurring task creation with specific timing. This matches our need exactly.
    '''

    # ✅ CORRECT: Call the existing skill directly
    result = await schedule_recurring_task(
        name="Team Meeting",
        description="Weekly team sync",
        frequency="weekly",
        weekday="MO",
        time="09:00"
    )
    return result

    # ❌ WRONG: Don't reimplement what already exists
    # task_handle = await primitives.tasks.update("Create weekly task...")
    # This would duplicate the skill's logic unnecessarily
"""


def get_library_function_composition_example() -> str:
    """Example: composing multiple library functions for a complex goal.

    Shows how to orchestrate existing skills when the goal requires multiple steps.
    """

    return """
# Example: Composing library functions for a complex workflow
async def main_plan():
    '''User goal: Find all contacts in New York and send them a meeting invite.

    ANALYSIS: The library has:
    - 'filter_contacts_by_location' for geographic filtering
    - 'send_bulk_notification' for batch messaging
    These can be composed to achieve the goal.
    '''

    # Step 1: Use existing skill to filter contacts
    ny_contacts = await filter_contacts_by_location(location="New York")

    # Step 2: Use existing skill to send notifications
    result = await send_bulk_notification(
        contact_list=ny_contacts,
        message="Team meeting tomorrow at 2pm",
        medium="email"
    )

    return f"Sent invites to {result['count']} contacts in New York"
"""


def get_library_function_adaptation_example() -> str:
    """Example: adapting an existing skill for a slightly different use case.

    Shows when to call a skill with different parameters vs reimplementing.
    """

    return """
# Example: Adapting a library function with parameters
async def main_plan():
    '''User goal: Get the phone number for the contact who works at Tesla.

    ANALYSIS: The library has 'find_contact_by_attribute' which can find contacts
    by any attribute. We can use it with employer="Tesla" to match our need.
    '''

    # ✅ CORRECT: Adapt the existing skill with appropriate parameters
    from pydantic import BaseModel

    class ContactResult(BaseModel):
        name: str
        phone: str | None
        employer: str | None

    ContactResult.model_rebuild()

    # The existing skill accepts flexible attribute filtering
    tesla_contact = await find_contact_by_attribute(
        attribute="employer",
        value="Tesla",
        response_format=ContactResult
    )

    return f"{tesla_contact.name}'s phone: {tesla_contact.phone or 'Not available'}"
"""


# ---------------------------------------------------------------------------
# 3. Computer Examples
# ---------------------------------------------------------------------------


def get_computer_navigation_example() -> str:
    """Example: navigate and extract data from a webpage."""

    return '''
# Example: Computer navigation and extraction
async def fetch_product_price(product_url: str) -> float:
    """Navigate to product page and extract price."""
    session = await primitives.computer.web.new_session()
    notify({"type": "progress", "message": f"Looking up pricing on {product_url}."})
    await session.navigate(product_url)

    from pydantic import BaseModel

    class ProductInfo(BaseModel):
        name: str
        price: float
        in_stock: bool

    ProductInfo.model_rebuild()

    info = await session.observe(
        "Extract product name, price, and stock status",
        response_format=ProductInfo
    )
    await session.stop()
    return info.price
'''


def get_computer_multistep_example() -> str:
    """Example: multi-step computer workflow with verification."""

    return '''
# Example: Multi-step computer workflow
async def complete_checkout(cart_items: list) -> str:
    """Complete e-commerce checkout flow."""
    session = await primitives.computer.web.new_session()
    notify({"type": "progress", "message": "Starting checkout for your order."})
    await session.navigate("https://shop.example.com/checkout")

    await session.act("Fill shipping address: 123 Main St, City, 12345")
    verification = await session.observe(
        "Is the shipping address '123 Main St, City, 12345' displayed?"
    )
    if "no" in verification.lower():
        raise ValueError("Shipping address verification failed")

    notify({"type": "progress", "message": "Placing the order now."})
    await session.act("Click 'Complete Order' button")
    confirmation = await session.observe("Extract order confirmation number")
    await session.stop()
    return f"Order placed: {confirmation}"
'''


def get_computer_screenshot_driven_example() -> str:
    """Example: screenshot-driven implementation using get_screenshot() + display().

    UI actions are guided by the current page state (captured via explicit screenshot calls).
    """

    return """
# Example: Screenshot-driven implementation
async def proceed_using_screenshot() -> str:
    session = await primitives.computer.web.new_session()
    notify({"type": "progress", "message": "Running through the setup wizard."})
    await session.navigate("https://example.com/setup")

    # Use get_screenshot() + display() to see the current screen state.
    # Prefer acting directly from that visual context, and only use observe
    # for structured extraction or when a precise, machine-checkable answer is required.
    display(await session.get_screenshot())
    await session.act("Click the 'Continue' button.")
    display(await session.get_screenshot())

    result = await session.observe("Confirm we reached the next step.")
    await session.stop()
    return result
"""


def get_computer_session_execution_example() -> str:
    """Example: web navigation and structured data extraction using session-based execution."""

    return """
**Example: Web Navigation and Structured Data Extraction**

*User Request*: "What is the main heading and the text of the first paragraph on playwright.dev?"

*Turn 1: Navigate to the website and view the page*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "The first step is to navigate to the website specified in the user\'s request, which is playwright.dev. I\'ll take a screenshot to see the page.",
          "code": "session = await primitives.computer.web.new_session()\\nnotify({\\"type\\": \\"progress\\", \\"message\\": \\"Opening playwright.dev to read the page.\\"})\\nawait session.navigate(\'https://playwright.dev/\')\\ndisplay(await session.get_screenshot())",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- stdout ---
    [screenshot image of the playwright.dev homepage]
    ```

*Turn 2: Observe the content using a Pydantic model*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "Great, I can see the page. Now I\'ll extract the heading and paragraph text into a structured object for clarity. I\'ll define a Pydantic model right here in the sandbox.",
          "code": "from pydantic import BaseModel, Field\\n\\nclass PageContent(BaseModel):\\n    heading: str = Field(description=\\"The main H1 heading of the page\\")\\n    first_paragraph: str = Field(description=\\"The text of the first paragraph under the heading\\")\\n\\nPageContent.model_rebuild()\\n\\npage_info = await session.observe(\\n    \\"Extract the main heading and the first paragraph.\\",\\n    response_format=PageContent\\n)\\nprint(page_info.model_dump_json(indent=2))\\nawait session.stop()",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- stdout ---
    {
      "heading": "Playwright enables reliable end-to-end testing for modern web apps.",
      "first_paragraph": "Playwright is an open-source framework for web testing and automation. It allows testing Chromium, Firefox and WebKit with a single API."
    }
    ```

*Turn 3: Provide the final answer*
* **Thought**: I have successfully extracted the information. I will now provide the final answer to the user without using any tools.
* **Final Answer (tool-less response)**:
    The main heading on playwright.dev is \'Playwright enables reliable end-to-end testing for modern web apps.\', and the first paragraph is \'Playwright is an open-source framework for web testing and automation. It allows testing Chromium, Firefox and WebKit with a single API.\'
"""


def get_computer_session_reattachment_example() -> str:
    """Example: reattach to an existing visible browser session by ID."""

    return """
# Example: Reusing an existing visible browser session
async def continue_existing_browser() -> str:
    # The current state or surrounding system may already tell you that
    # Web session 0 is active. Reattach to it instead of opening a duplicate.
    session = primitives.computer.web.get_session(0)

    # Verify the current state before acting.
    display(await session.get_screenshot())

    await session.act("Click the Continue button on the current page.")
    confirmation = await session.observe("Confirm the next step is now visible.")
    return confirmation
"""


def get_computer_stateful_workflow_example() -> str:
    """Example: stateful computation with helper functions persisting across turns."""

    return """
**Example: Stateful Computation and Helper Functions**

*User Request*: "On `fakestore.example.com`, find the average price of all products that cost less than $100."

*Turn 1: Define a helper function and extract all product data*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "This is a multi-step task. First, I\'ll extract all products. I know I\'ll need to parse prices that might be strings (e.g., \'$25.99\'), so I\'ll define a helper function to clean them. This function will persist in the sandbox for later.",
          "code": "session = await primitives.computer.web.new_session()\\nimport re\\nfrom pydantic import BaseModel, Field\\nfrom typing import List\\n\\ndef parse_price(price_str: str) -> float:\\n    nums = re.findall(r\'[\\\\d.]+\', price_str)\\n    return float(nums[0]) if nums else 0.0\\n\\nclass Product(BaseModel):\\n    name: str\\n    price_text: str = Field(alias=\\"price\\")\\n\\nclass ProductList(BaseModel):\\n    products: List[Product]\\n\\nProductList.model_rebuild()\\n\\nnotify({\\"type\\": \\"progress\\", \\"message\\": \\"Extracting product data from the store page.\\"})\\nall_products_data = await session.observe(\\n    \\"Extract all products with their name and price text\\",\\n    response_format=ProductList\\n)\\nprint(f\\"Extracted {len(all_products_data.products)} products.\\")\\nawait session.stop()",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- STDOUT ---
    Extracted 20 products.
    ```

*Turn 2: Use the helper function and the stored variable to compute the average*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "I have the product data in the `all_products_data` variable and my `parse_price` function is defined. Now I can perform the calculation in pure Python.",
          "code": "prices_under_100 = []\\nfor product in all_products_data.products:\\n    price = parse_price(product.price_text)\\n    if price < 100.0:\\n        prices_under_100.append(price)\\n\\nif prices_under_100:\\n    average = sum(prices_under_100) / len(prices_under_100)\\n    result_text = f\\"The average price of products under $100 is ${average:.2f}.\\"\\nelse:\\n    result_text = \\"No products found under $100.\\"\\nprint(result_text)",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- STDOUT ---
    The average price of products under $100 is $42.75.
    ```
* **Final Answer (tool-less)**: The average price of products under $100 on the site is $42.75.
"""


def get_computer_interactive_workflow_example() -> str:
    """Example: multi-step web form workflow with handle-based state manager interaction."""

    return """
**Example: Web Research and State Manager Persistence**

*User Request*: "Find the support email on example.com and save it to our knowledge base."

*Turn 1: Navigate to the website and view the page*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "I\'ll navigate to the website to find the support email and take a screenshot to see the page.",
          "code": "session = await primitives.computer.web.new_session()\\nnotify({\\"type\\": \\"progress\\", \\"message\\": \\"Looking up the support email on example.com.\\"})\\nawait session.navigate(\'https://example.com/contact\')\\ndisplay(await session.get_screenshot())",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- stdout ---
    [screenshot image of the example.com contact page]
    ```

*Turn 2: Extract the support email using structured observation*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_code",
        "arguments": {
          "thought": "I can see the contact page. I\'ll extract the support email using a Pydantic model for reliable structured extraction.",
          "code": "from pydantic import BaseModel\\n\\nclass ContactInfo(BaseModel):\\n    support_email: str\\n    phone: str | None = None\\n\\nContactInfo.model_rebuild()\\n\\ninfo = await session.observe(\\n    \\"Extract the support email address and phone number from the contact page.\\",\\n    response_format=ContactInfo\\n)\\nprint(f\\"Support email: {info.support_email}\\")\\nawait session.stop()",
          "language": "python",
          "state_mode": "stateful"
        }
      }]
    }
    ```
* **Observation**:
    ```text
    --- STDOUT ---
    Support email: help@example.com
    ```

*Turn 3: Persist the extracted info to the knowledge base (single primitive → execute_function)*
* **Tool Call**:
    ```json
    {
      "tool_calls": [{
        "name": "execute_function",
        "arguments": {
          "thought": "I have the support email (help@example.com). A single knowledge update is best done via execute_function so the outer loop retains the handle.",
          "function_name": "primitives.knowledge.update",
          "call_kwargs": {"text": "Store that Example Corp\'s support email is help@example.com"}
        }
      }]
    }
    ```
* **Observation**:
    ```text
    Successfully stored: Example Corp support email is help@example.com
    ```
* **Final Answer (tool-less)**: I found the support email on example.com — it\'s help@example.com — and saved it to the knowledge base.
"""


# ---------------------------------------------------------------------------
# 4. Primitives Examples (State Managers)
# ---------------------------------------------------------------------------


def get_primitives_contact_ask_example() -> str:
    """Example: read-only contact query."""

    return '''
# Example: Read-only contact query
async def find_contact_email(name: str) -> str:
    """Find a contact's email address by name."""
    handle = await primitives.contacts.ask(f"What is {name}'s email address?")
    answer = await handle.result()

    # Extract email from natural language answer
    # (In practice, use response_format for structured output)
    return answer
'''


def get_primitives_contact_update_example() -> str:
    """Example: contact mutation."""

    return '''
# Example: Contact mutation
async def update_contact_phone(email: str, phone: str) -> str:
    """Update a contact's phone number."""
    instruction = f"Update the contact with email {email}: set phone to {phone}"
    handle = await primitives.contacts.update(instruction)
    result = await handle.result()
    return result
'''


def get_primitives_cross_manager_example() -> str:
    """Example: cross-manager workflow (KnowledgeManager → ContactManager)."""

    return '''
# Example: Cross-manager workflow
async def find_employee_count_for_contact(contact_name: str) -> int:
    """Find how many employees work at a contact's company.

    This demonstrates cross-manager integration:
    1. KnowledgeManager.ask internally calls ContactManager.ask to find employer
    2. KnowledgeManager.ask then queries its own tables for employee count
    """
    # Single high-level query; KM handles cross-manager coordination internally
    handle = await primitives.knowledge.ask(
        f"How many employees are at the company {contact_name} works at?"
    )
    answer = await handle.result()

    # Extract count from natural language answer
    # (In practice, use response_format for structured output)
    import re
    match = re.search(r'(\\d+)', answer)
    return int(match.group(1)) if match else 0
'''


def get_primitives_task_execute_example() -> str:
    """Example: task execution with steering."""

    return '''
# Example: Task execution with steering
async def execute_task_by_description_with_guidance(description: str) -> str:
    """Find and execute a task by description, with mid-flight guidance."""
    from pydantic import BaseModel

    class TaskIdResult(BaseModel):
        task_id: int
        task_name: str

    TaskIdResult.model_rebuild()

    notify({"type": "progress", "message": f"Looking up and running the task: {description}."})

    # Step 1: Find the task_id using structured output from `ask(...)`.
    lookup_handle = await primitives.tasks.ask(
        f"Find the task that best matches: {description}. Return the task_id and name.",
        response_format=TaskIdResult,
    )
    task_info = await lookup_handle.result()

    # Step 2: Execute using the task_id (returns a steerable ActiveQueue handle).
    handle = await primitives.tasks.execute(task_id=task_info.task_id)

    # Inject guidance early in execution
    await handle.interject("Provide a progress update after each major step.")

    # Query status mid-execution
    status_handle = await handle.ask("What is the current status?")
    status = await status_handle.result()

    # Stop early if needed
    if "error" in status.lower():
        await handle.stop(reason="Detected error in status")

    # Wait for completion
    result = await handle.result()
    return result
'''


def get_primitives_dynamic_methods_example() -> str:
    """Example: using dynamic handle methods."""

    return '''
# Example: Dynamic handle methods (append_to_queue)
async def execute_task_and_append(task_a_id: int, task_b_id: int) -> str:
    """Execute a task and append another to its queue.

    TaskScheduler.execute returns an ActiveQueue handle that exposes
    a dynamic method: append_to_queue_<tool>_<id>(task_id=...)
    """
    notify({"type": "progress", "message": f"Running task {task_a_id} and appending task {task_b_id} to the queue."})

    # Start task A
    handle = await primitives.tasks.execute(task_id=task_a_id)

    # The handle exposes a dynamic append method
    # (exact name depends on handle instance; use introspection or ask)
    # For this example, assume we know the method name pattern

    # Append task B to the queue while A is running
    # Note: In practice, the LLM discovers this method via tool introspection
    await handle.append_to_queue(task_id=task_b_id)

    # Wait for completion (both tasks will execute in order)
    result = await handle.result()
    return result
'''


def get_primitives_files_describe_example() -> str:
    """Example: discovering file storage layout via `primitives.files.describe(...)`."""

    return '''
# Example: Discover file storage layout
async def discover_file_structure(file_path: str) -> dict:
    """Discover what tables/contexts are available in a file."""
    storage = await primitives.files.describe(file_path=file_path)

    # storage contains:
    # - indexed_exists: bool (whether file has been indexed)
    # - has_tables: bool (whether tables were extracted)
    # - tables: list of table info with context_path, name, description

    if storage.has_tables:
        for table in storage.tables:
            print(f"Table: {table.name}")
            print(f"  Context: {table.context_path}")
            print(f"  Description: {table.description}")
    return storage
'''


def get_primitives_files_reduce_example() -> str:
    """Example: aggregating data via `primitives.files.reduce(...)`."""

    return '''
# Example: Aggregate data from a file table
async def count_records_by_category(table_context: str) -> dict:
    """Count records grouped by a category column."""
    result = await primitives.files.reduce(
        context=table_context,
        metric="count",
        columns="id",
        group_by="category",
        filter="status == 'active'",
    )
    return result
'''


def get_primitives_files_filter_example() -> str:
    """Example: filtering rows via `primitives.files.filter_files(...)`."""

    return '''
# Example: Filter rows from a file table
async def get_recent_records(table_context: str) -> list:
    """Get recent records matching a filter."""
    rows = await primitives.files.filter_files(
        context=table_context,
        filter="created_date > '2024-01-01'",
        columns=["id", "name", "created_date", "status"],
        limit=50,
    )
    return rows
'''


def get_primitives_files_search_example() -> str:
    """Example: semantic search via `primitives.files.search_files(...)`."""

    return '''
# Example: Semantic search over file data
async def search_for_topic(table_context: str, query: str) -> list:
    """Search for records semantically matching a query."""
    hits = await primitives.files.search_files(
        context=table_context,
        references={"description": query},
        limit=10,
    )
    return hits
'''


def get_primitives_files_visualize_example() -> str:
    """Example: generating charts via `primitives.files.visualize(...)`."""

    return '''
# Example: Generate a chart from file data
async def plot_distribution(table_context: str) -> str:
    """Generate a bar chart showing distribution by category."""
    result = await primitives.files.visualize(
        tables=table_context,
        plot_type="bar",
        x_axis="category",
        y_axis="amount",
        metric="sum",
        title="Amount by Category",
    )
    return result.get("url") if isinstance(result, dict) else result
'''


def get_primitives_files_render_extract_example() -> str:
    """Example: visual render-first extraction from Excel and PDF files."""

    return '''
# Example: Extract structured data from Excel and PDF documents
# The render-first workflow is the most robust approach for document
# extraction — it handles scanned PDFs, complex layouts, merged cells,
# and formatting cues that text-based parsing misses.
import openpyxl

async def extract_financials_from_documents(directory: str) -> dict:
    """Extract financial data from Excel and PDF files in a directory."""
    import os

    results = {}

    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)

        if filename.endswith((".xlsx", ".xls")):
            # --- Excel: render full sheet first, then zoom in ---
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.sheet_state != "visible":
                    continue

                # Step 1: Global view — render the full sheet
                overview = await primitives.files.render_excel_sheet(sheet)
                display(overview)

                # Step 2: Zoom into a specific area of interest
                detail = await primitives.files.render_excel_sheet(
                    sheet, cell_range="A1:H30",
                )
                display(detail)

                # Step 3: Read exact cell values via openpyxl
                for row in sheet.iter_rows(min_row=2, max_col=8, values_only=False):
                    label = row[0].value
                    value = row[1].value
                    if label and value is not None:
                        results[label] = {
                            "value": value,
                            "source": f"{sheet_name}!{row[0].coordinate}",
                        }

        elif filename.endswith(".pdf"):
            # --- PDF: render 2-3 pages at a time ---
            import pymupdf

            doc = pymupdf.open(filepath)
            total_pages = len(doc)
            for page_num in range(0, total_pages, 2):
                img = await primitives.files.render_pdf(filepath, page=page_num)
                display(img)
                if page_num + 1 < total_pages:
                    img2 = await primitives.files.render_pdf(
                        filepath, page=page_num + 1,
                    )
                    display(img2)
            doc.close()

    return results
'''


def get_primitives_web_ask_example() -> str:
    """Example: web research query via `primitives.web.ask(...)`."""

    return '''
# Example: Web research query
async def research_latest_news() -> str:
    """Ask the WebSearcher for time-sensitive info."""
    handle = await primitives.web.ask("What are the major world news headlines this week?")
    result = await handle.result()
    return result
'''


# ---------------------------------------------------------------------------
# 5. Mixed Examples (Computer + Primitives)
# ---------------------------------------------------------------------------


def get_mixed_browse_persist_example() -> str:
    """Example: browse for data and persist via state managers."""

    return '''
# Example: Browse and persist workflow
async def scrape_and_save_contact(linkedin_url: str) -> str:
    """Scrape contact info from LinkedIn and save to ContactManager."""
    session = await primitives.computer.web.new_session()
    notify({"type": "progress", "message": "Looking up the profile and saving contact details."})
    await session.navigate(linkedin_url)

    from pydantic import BaseModel
    from typing import Optional

    class LinkedInProfile(BaseModel):
        name: str
        email: Optional[str] = None
        company: str

    LinkedInProfile.model_rebuild()

    profile = await session.observe(
        "Extract name, email, and current company from profile",
        response_format=LinkedInProfile
    )
    await session.stop()

    instruction = f"Create contact: {profile.name}, email {profile.email}, employer {profile.company}"
    handle = await primitives.contacts.update(instruction)
    result = await handle.result()
    return f"Saved contact: {result}"
'''


def get_mixed_concurrent_example() -> str:
    """Example: concurrent computer and state manager operations."""

    return '''
# Example: Concurrent computer + state manager operations
import asyncio

async def gather_contact_info_concurrently(name: str, company_url: str) -> dict:
    """Gather contact info from multiple sources concurrently."""
    notify({"type": "progress", "message": f"Looking up {name}'s details and company info."})
    contact_handle = await primitives.contacts.ask(f"Find {name}'s email and phone")

    async def fetch_company_info():
        session = await primitives.computer.web.new_session()
        await session.navigate(company_url)
        result = await session.observe("Extract company size and industry")
        await session.stop()
        return result

    contact_result, company_info = await asyncio.gather(
        contact_handle.result(),
        fetch_company_info()
    )
    return {
        "contact": contact_result,
        "company": company_info
    }
'''


def get_mixed_interjection_routing_example() -> str:
    """Example: routing interjections to in-flight handles."""

    return """
# Example: Interjection routing to in-flight handles
async def search_multiple_sources_with_correction(query: str) -> dict:
    # Start multiple concurrent searches
    notify({
        "type": "progress",
        "message": "Starting parallel contact and transcript searches.",
        "step": 1,
        "total": 2
    })
    contact_handle = await primitives.contacts.ask(query)
    transcript_handle = await primitives.transcripts.ask(query)

    # If user interjects with clarification (e.g., "I meant David Smith"),
    # the Actor's pane can broadcast the interjection to all in-flight handles
    # via: pane.broadcast_interject("User clarified: David Smith", filter=...)

    # Wait for results (interjections handled by pane)
    contact_result = await contact_handle.result()
    transcript_result = await transcript_handle.result()
    return {
        "contacts": contact_result,
        "transcripts": transcript_result
    }
"""


# ---------------------------------------------------------------------------
# 5b. Interjection Examples (Routing vs Patching)
# ---------------------------------------------------------------------------


def get_interjection_routing_only_examples() -> str:
    """Examples: routing-only interjections (no patches / no cache invalidation).

    These examples are intentionally written with *natural user language* (no "broadcast",
    no internal references like "main_plan"), and demonstrate the preferred output shape
    when the interjection can be satisfied by steering *already in-flight* handles.
    """

    return """
### Interjection Examples: Routing-Only (NO PATCHES)

**Rule of thumb:** If the user’s interjection can be satisfied by *steering what’s already running*
(tone/formatting/scope preferences for current outputs), prefer **routing-only**:
- Do **NOT** emit `patches`
- Do **NOT** emit `cache`
- Use `routing_action="broadcast_filtered"` (or `"targeted"` if the user singled out a specific handle)

#### Example A — Targeted tone adjustment (only the relevant handle)
- **Context**: There is an in-flight `primitives.contacts.ask(...)` that is just listing contacts, and an in-flight
  `primitives.transcripts.ask(...)` that is generating outreach tone/greeting guidance.
- **User**: “Actually—let’s make the outreach much more casual and friendly. Use ‘Hey <FirstName>,’ and don’t use last names or ‘Dear’.”
- **Correct JSON** (route only to the *relevant* in-flight handle):
```json
{
  "action": "modify_task",
  "reason": "This is a preference update for what is currently being produced by an in-flight handle. Routing is sufficient; patching would be unnecessarily disruptive.",
  "routing_action": "targeted",
  "target_handle_ids": ["<transcripts_handle_id_from_pane_snapshot>"],
  "routed_message": "Make the outreach much more casual and friendly. Use greeting 'Hey <FirstName>,' (first name only). Avoid last names and do not use 'Dear'."
}
```

#### Example D — Targeted refinement of an in-flight lookup (ranking / tie-breaker)
- **Context**: There is an in-flight `primitives.tasks.update(...)` creating a task, and an in-flight
  `primitives.contacts.ask(...)` identifying an assignee. The user adds a *selection preference* for the in-flight lookup.
- **User**: “For the Berlin contact lookup, prioritize anyone who has ‘manager’ or ‘lead’ in their role.”
- **Correct JSON** (route only; do NOT patch/restart):
```json
{
  "action": "modify_task",
  "reason": "This is a refinement for how the in-flight contact lookup should choose/format its answer. Steering the running contacts handle is sufficient; patching would unnecessarily cancel/restart execution.",
  "routing_action": "targeted",
  "target_handle_ids": ["<contacts_handle_id_from_pane_snapshot>"],
  "routed_message": "For the Berlin contact lookup: prioritize anyone whose role/title contains 'manager' or 'lead'. If multiple match, pick the most senior."
}
```

#### Example E — Upstream scope correction BEFORE downstream steps consume the result (staged workflow)
- **Context**: There is an in-flight `primitives.contacts.ask(...)` producing a contact set. The plan has *not yet awaited* that handle’s `.result()`,
  and downstream steps (e.g., transcripts/knowledge) will be spawned *based on* the returned contacts.
- **User**: “Actually, for this quarterly sync, let’s focus only on the Berlin office. When you list the active contacts, include only Berlin contacts.”
- **Correct JSON** (route only to the upstream handle; downstream will naturally use the corrected `contacts_result`):
```json
{
  "action": "modify_task",
  "reason": "The correction applies to an already in-flight upstream lookup whose result will feed downstream steps. Steering the in-flight contacts handle is sufficient; patching/restarting is unnecessarily disruptive.",
  "routing_action": "targeted",
  "target_handle_ids": ["<contacts_handle_id_from_pane_snapshot>"],
  "routed_message": "Scope update for the contact list currently in progress: include ONLY Berlin office contacts (names + office + email)."
}
```

#### Example B — Scope filter (“focus only on Q4”)
- **User**: “One change: for the summaries you’re producing right now, focus only on Q4 2024 and ignore Q3.”
- **Correct JSON**:
```json
{
  "action": "modify_task",
  "reason": "This is a scope constraint for the current in-flight summaries; route it to the running handles instead of patching plan code.",
  "routing_action": "broadcast_filtered",
  "broadcast_filter": {
    "capabilities": ["interjectable"],
    "origin_tool_prefixes": ["primitives.contacts", "primitives.transcripts", "primitives.knowledge"]
  },
  "routed_message": "For the summaries currently in progress: focus ONLY on Q4 2024 (Oct/Nov/Dec). Ignore Q3 entirely and do not mention it."
}
```

#### Example C — Conciseness (“keep it short”)
- **User**: “Keep everything you’re working on right now concise. Only the essentials—no quotes, no extra detail.”
- **Correct JSON**:
```json
{
  "action": "modify_task",
  "reason": "This is a formatting preference for outputs of in-flight handles; routing-only is the correct and least disruptive approach.",
  "routing_action": "broadcast_filtered",
  "broadcast_filter": { "capabilities": ["interjectable"] },
  "routed_message": "Be concise and to the point. Only include essential info. No quotes. No extra detail."
}
```

#### Anti-pattern (avoid circular “patch so replay is consistent”)
- **User**: “Keep the summaries you’re producing right now concise.”
- **Wrong reasoning**: “I should patch the plan prompts so that if we replay, it stays concise.”
- **Why wrong**: Adding `patches` triggers cancellation/restart and cache invalidation. If routing is sufficient,
  don’t force a restart “for replay consistency”.
""".strip()


# ---------------------------------------------------------------------------
# 6. Verification Examples (Prompt Builders)
# ---------------------------------------------------------------------------


def get_verification_ok_example() -> str:
    """Verification example: clear success (environment-agnostic).

    Shows how to map trace + evidence → decision.
    """
    return """
### Verification Example: ok (core)
- **Goal**: "Add Alice to my contacts."
- **Intent**: `create_contact(name='Alice', email='alice@corp.com')`
- **Agent Trace**:
  - `REASONING: I will create the contact and confirm the result.`
  - `notify({'type': 'progress', 'message': 'Updating the contact now.'})`
  - `tool_call: primitives.contacts.update(...)`
  - `✓ done`
- **Evidence**: Return value indicates success.
- **Decision**:
```json
{"status": "ok", "reason": "Trace shows correct action; return value confirms contact was created, so this step meaningfully advances the goal."}
```
""".strip()


def get_verification_ambiguous_example() -> str:
    """Verification example: ambiguity → request clarification (environment-agnostic)."""
    return """
### Verification Example: request_clarification (core)
- **Goal**: "Invite Bob to the calendar event."
- **Intent**: `invite_attendee(name='Bob')`
- **Agent Trace**:
  - `REASONING: I found two Bobs; I chose Bob S. and proceeded.`
  - `notify({'type': 'progress', 'message': 'Looking up matching contacts.'})`
  - `tool_call: primitives.contacts.ask(...)`
- **Evidence**: The trace reveals an unverified assumption (which Bob).
- **Decision**:
```json
{"status": "request_clarification", "reason": "The trace shows an ambiguous choice between two valid contacts; outcome cannot be verified as correct.", "clarification_question": "I found two contacts named Bob. Which one should I invite (e.g. Bob Smith or Bob Stone)?"}
```
""".strip()


def get_verification_tactical_failure_example() -> str:
    """Verification example: tactical failure → reimplement locally (environment-agnostic)."""
    return """
### Verification Example: reimplement_local (core)
- **Goal**: "Extract the support email from the page."
- **Intent**: `extract_support_email()`
- **Agent Trace**:
  - `REASONING: I see an email-like string; I will return it.`
  - `returned: 'support@example.com'`
- **Evidence**: Screenshot/visible text actually shows `help@company.com` as the support email.
- **Decision**:
```json
{"status": "reimplement_local", "reason": "The returned value does not match evidence. Root cause: extraction logic picked the wrong email string. Fix strategy: re-implement to search for the labeled 'Support' section and extract the email adjacent to it."}
```
""".strip()


def get_verification_strategic_failure_example() -> str:
    """Verification example: strategic failure → replan parent (environment-agnostic)."""
    return """
### Verification Example: replan_parent (core)
- **Goal**: "Apply a price filter under $500."
- **Intent**: `apply_price_filter(max_price=500)`
- **Agent Trace**:
  - `REASONING: I cannot find any filter controls on this page.`
  - `✓ done (no action possible)`
- **Evidence**: Screenshot shows a landing page without product list or filters.
- **Decision**:
```json
{"status": "replan_parent", "reason": "The function is impossible because the parent failed to navigate to a product listing page with filters. Fix strategy: update the parent to navigate into the catalog/results page before calling this function."}
```
""".strip()


def get_computer_verification_extraction_example() -> str:
    """Verification example: extraction (computer environment)."""
    return """
### Computer Verification Example: extraction
- **Goal**: "Find the product price."
- **Intent**: `extract_price()`
- **Agent Trace**:
  - `notify({'type': 'progress', 'message': 'Reading the page to capture the current price.'})`
  - `session.observe('Extract price')`
  - `returned: '$199'`
- **Evidence**: Screenshot shows the price as `$299` (the `$199` is the crossed-out old price).
- **Decision**:
```json
{"status": "reimplement_local", "reason": "Tactical mismatch: trace executed, but evidence contradicts return value. Fix strategy: re-extract the *current* price (non-struck, highlighted) and verify against the screenshot."}
```
""".strip()


def get_computer_verification_multistep_example() -> str:
    """Verification example: multi-step success (computer environment)."""
    return """
### Computer Verification Example: multistep ok
- **Goal**: "Submit the signup form."
- **Intent**: `submit_signup(email='a@corp.com')`
- **Agent Trace**:
  - `notify({'type': 'progress', 'message': 'Entering signup details.'})`
  - `session.act('Type email…')`
  - `notify({'type': 'progress', 'message': 'Submitting the signup form.'})`
  - `session.act('Click Sign up')`
  - `✓ done`
- **Evidence**: Screenshot shows a confirmation banner 'Thanks for signing up'.
- **Decision**:
```json
{"status": "ok", "reason": "Trace shows correct actions and screenshot confirms successful submission, so the function advanced the overall goal."}
```
""".strip()


def get_primitives_verification_contact_update_example() -> str:
    """Verification example: primitives update success (primitives environment)."""
    return """
### Primitives Verification Example: contact update ok
- **Goal**: "Update Carol's phone number."
- **Intent**: `update_phone(email='carol@corp.com', phone='555-0101')`
- **Agent Trace**:
  - `notify({'type': 'progress', 'message': 'Applying the requested contact update.'})`
  - `primitives.contacts.update(...)`
  - `✓ done`
- **Evidence**: Return value indicates the contact was updated successfully.
- **Decision**:
```json
{"status": "ok", "reason": "The update call completed and the return value indicates success; this is sufficient evidence for a state-manager mutation."}
```
""".strip()


def get_primitives_verification_cross_manager_example() -> str:
    """Verification example: primitives cross-manager strategic failure (primitives environment)."""
    return """
### Primitives Verification Example: cross-manager replan_parent
- **Goal**: "Add the CEO of Acme to contacts."
- **Intent**: `create_ceo_contact(company='Acme')`
- **Agent Trace**:
  - `REASONING: I queried the company but found no CEO field.`
  - `notify({'type': 'progress', 'message': 'Searching company details for CEO information.'})`
  - `primitives.knowledge.ask(...)`
  - `✓ done (insufficient data)`
- **Evidence**: Return value indicates missing CEO data; nothing to create.
- **Decision**:
```json
{"status": "replan_parent", "reason": "Strategic failure: the parent plan assumed CEO data existed. Fix strategy: replan to first find a reliable source for the CEO (or ask user) before calling contact creation."}
```
""".strip()


def get_mixed_verification_browse_persist_example() -> str:
    """Verification example: mixed browse + persist success (computer + primitives)."""
    return """
### Mixed Verification Example: browse + persist ok
- **Goal**: "Find support email on the site and save it to Knowledge."
- **Intent**: `scrape_support_email_and_save()`
- **Agent Trace**:
  - `notify({'type': 'progress', 'message': 'Reading the site for the support email.'})`
  - `session.observe('Extract support email') -> 'help@company.com'`
  - `notify({'type': 'progress', 'message': 'Saving the support email to shared knowledge.'})`
  - `primitives.knowledge.update('Save support_email=help@company.com')`
  - `✓ done`
- **Evidence**: Screenshot shows the extracted email; update return value indicates persistence succeeded.
- **Decision**:
```json
{"status": "ok", "reason": "The computer evidence matches the extracted email, and the state-manager update confirms it was saved."}
```
""".strip()


def get_verification_examples_for_environments(
    has_computer: bool,
    has_primitives: bool,
) -> str:
    """Return verification examples appropriate for the given environment combination."""
    sections: list[str] = []

    # Core verification examples are always useful (environment-agnostic).
    sections.append(
        "### Core Verification Examples (Environment-Agnostic)\n"
        + "\n\n".join(
            [
                get_verification_ok_example().strip(),
                get_verification_ambiguous_example().strip(),
            ],
        ),
    )

    if has_computer:
        sections.append(
            "### Computer Verification Examples\n"
            + "\n\n".join(
                [
                    get_computer_verification_extraction_example().strip(),
                    get_computer_verification_multistep_example().strip(),
                ],
            ),
        )

    if has_primitives:
        sections.append(
            "### Primitives Verification Examples\n"
            + "\n\n".join(
                [
                    get_primitives_verification_contact_update_example().strip(),
                    get_primitives_verification_cross_manager_example().strip(),
                ],
            ),
        )

    if has_computer and has_primitives:
        sections.append(
            "### Mixed Verification Examples\n"
            + get_mixed_verification_browse_persist_example().strip(),
        )

    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# 7. Example Function Map (for ToolSurfaceRegistry)
# ---------------------------------------------------------------------------


def get_example_function_map() -> dict[str, callable]:
    """Get a mapping of example function names to their callables.

    This is used by ToolSurfaceRegistry.prompt_examples() to generate
    examples for exposed managers.
    """
    return {
        # Contacts
        "get_primitives_contact_ask_example": get_primitives_contact_ask_example,
        "get_primitives_contact_update_example": get_primitives_contact_update_example,
        # Tasks
        "get_primitives_task_execute_example": get_primitives_task_execute_example,
        "get_primitives_dynamic_methods_example": get_primitives_dynamic_methods_example,
        # Knowledge
        "get_primitives_knowledge_ask_example": get_primitives_cross_manager_example,
        "get_primitives_knowledge_update_example": lambda: "",  # placeholder
        # Transcripts
        "get_primitives_transcript_ask_example": lambda: "",  # placeholder
        # Files (using real FileManager primitives)
        "get_primitives_files_render_extract_example": get_primitives_files_render_extract_example,
        "get_primitives_files_describe_example": get_primitives_files_describe_example,
        "get_primitives_files_reduce_example": get_primitives_files_reduce_example,
        "get_primitives_files_filter_example": get_primitives_files_filter_example,
        "get_primitives_files_search_example": get_primitives_files_search_example,
        "get_primitives_files_visualize_example": get_primitives_files_visualize_example,
        # Web
        "get_primitives_web_ask_example": get_primitives_web_ask_example,
        # Secrets
        "get_primitives_secrets_ask_example": lambda: "",  # placeholder
        "get_primitives_secrets_update_example": lambda: "",  # placeholder
        # Data
        "get_primitives_data_filter_example": lambda: "",  # placeholder
        "get_primitives_data_reduce_example": lambda: "",  # placeholder
    }


# ---------------------------------------------------------------------------
# 8. Discovery-First Pattern Examples (for CodeActActor)
# ---------------------------------------------------------------------------


def get_discovery_first_pattern_example() -> str:
    """Example: prioritizing pre-saved functions and guidance via discovery tools (CodeAct style)."""

    return r"""
# ✅ PATTERN: Discovery-First Workflow (CodeActActor)
# If FunctionManager and GuidanceManager tools are available, ALWAYS search both
# for existing functions and guidance BEFORE writing custom logic with raw primitives.
#
# Step 1 (JSON TOOL CALL): search for an existing function
# {
#   "name": "FunctionManager_search_functions",
#   "arguments": {"query": "contacts prefer phone", "n": 5}
# }
#
# Step 2 (JSON TOOL CALL): execute the discovered function
# {
#   "name": "execute_code",
#   "arguments": {
#     "language": "python",
#     "state_mode": "stateless",
#     "code": "result = await ask_contacts_question('Which of our contacts prefers phone contact?')\nprint(result)"
#   }
# }
#
# FunctionManager-discovered functions are available in all execute_code calls
# (both stateful and stateless). Use stateful only when you need intermediate
# variables to persist across calls.
#
# If no function exists, THEN fall back to composing with primitives directly in Python.
# For multi-step compositions, emit `notify({...})` at meaningful milestones.
# For a single primitive call, use `execute_function` — no notify needed.
"""


def get_discovery_first_anti_pattern_example() -> str:
    """Anti-pattern: skipping FunctionManager/GuidanceManager search when they exist (CodeAct style)."""

    return r"""
# ❌ ANTI-PATTERN #1: Skipping FunctionManager when it's available
#
# DON'T do this:
#   - immediately call raw primitives
#   - re-implement logic that likely exists as a stored function
#
# Example (bad):
#   handle = await primitives.contacts.ask("Which contacts prefer phone?")
#   result = await handle.result()
#
# ✅ CORRECT:
#   1) Call FunctionManager_search_functions(...) as a JSON tool call
#   2) Call execute_code and invoke the discovered function (any state_mode works)
"""


def get_function_parameter_exploration_example() -> str:
    """Example: reading function metadata before calling (CodeAct style)."""

    return r"""
# ✅ PATTERN: Read function signatures before calling
#
# FunctionManager tool results include:
# - name
# - argspec
# - docstring
#
# Use that to pick the right parameters before calling the injected function in Python.
#
# Example:
# 1) JSON tool call:
#    FunctionManager_search_functions(query="update guidance runbook", n=5)
#
# 2) Inspect returned `argspec`/docstring (mentally), then in Python:
#    result = await update_guidance("Create a runbook titled 'Runbook: DB Failover' ...")
#    print(result)
"""


def get_code_act_discovery_first_examples() -> str:
    """Get discovery-first examples for CodeActActor."""
    examples = [
        get_discovery_first_pattern_example().strip(),
        get_discovery_first_anti_pattern_example().strip(),
        get_function_parameter_exploration_example().strip(),
    ]
    return "\n\n".join(examples)


# ---------------------------------------------------------------------------
# 8b. Multi-language + multi-session execution examples (for CodeActActor)
# ---------------------------------------------------------------------------


def get_code_act_session_examples() -> str:
    """Examples: using execute_code with sessions across Python + shell.

    These examples are written in JSON tool-call form (not Python), because
    `execute_code` is itself a tool call.
    """

    return r"""
### Multi-Language + Multi-Session Execution (CodeActActor)

**Key idea:** Use `execute_code` for *everything* (Python + shell), and use sessions
to preserve state across multiple tool calls.

> **Note**: FunctionManager-discovered functions are available in all `execute_code` calls regardless of `state_mode`.

#### Example A — Stateful shell session for repo navigation
```json
{
  "tool_calls": [{
    "name": "execute_code",
    "arguments": {
      "thought": "Create or reuse a persistent bash session for navigating the repo.",
      "language": "bash",
      "state_mode": "stateful",
      "session_name": "repo_nav",
      "code": "pwd && ls"
    }
  }]
}
```

#### Example B — Continue the same shell session (cd persists)
```json
{
  "tool_calls": [{
    "name": "execute_code",
    "arguments": {
      "thought": "Continue in the same session so cwd is preserved.",
      "language": "bash",
      "state_mode": "stateful",
      "session_name": "repo_nav",
      "code": "cd unity && ls && git status -sb"
    }
  }]
}
```

#### Example C — Stateful Python session for iterative work
```json
{
  "tool_calls": [{
    "name": "execute_code",
    "arguments": {
      "thought": "Use a stateful Python session so variables persist between calls.",
      "language": "python",
      "state_mode": "stateful",
      "session_name": "analysis",
      "code": "x = 41\nx += 1\nprint(x)"
    }
  }]
}
```

#### Example D — Session discovery and inspection
```json
{
  "tool_calls": [
    { "name": "list_sessions", "arguments": {} },
    {
      "name": "inspect_state",
      "arguments": {
        "detail": "names",
        "session_name": "analysis",
        "language": "python"
      }
    }
  ]
}
```

#### Example E — Stateless execution (no persistence)
```json
{
  "tool_calls": [{
    "name": "execute_code",
    "arguments": {
      "thought": "Run a one-off command with no session/persistence.",
      "language": "bash",
      "state_mode": "stateless",
      "code": "echo hello"
    }
  }]
}
```
""".strip()


# ---------------------------------------------------------------------------
# 9. Helper Functions for Prompt Composition
# ---------------------------------------------------------------------------


def get_core_pattern_examples() -> str:
    """Get all core pattern examples (environment-agnostic)."""

    examples = [
        get_library_function_reuse_example().strip(),
        get_library_function_composition_example().strip(),
        get_library_function_adaptation_example().strip(),
        get_confidence_based_stubbing_example().strip(),
        get_structured_output_example().strip(),
        get_error_handling_example().strip(),
        get_handle_steering_example().strip(),
        get_clarification_example().strip(),
        get_notify_web_search_example().strip(),
        get_notify_multistep_workflow_example().strip(),
        get_notify_long_running_example().strip(),
    ]
    return "\n\n".join(examples)


def get_code_act_pattern_examples() -> str:
    """Get core pattern examples relevant to CodeActActor.

    Includes error handling, clarification patterns, and discovery-first workflow
    that complement the primitives examples.
    """

    examples = [
        get_error_handling_example().strip(),
        get_handle_mode_selection_example().strip(),
        get_clarification_example().strip(),
        get_notify_web_search_example().strip(),
        get_notify_multistep_workflow_example().strip(),
        get_notify_long_running_example().strip(),
    ]
    return "\n\n".join(examples)


def get_computer_examples() -> str:
    """Get all computer-specific examples."""

    examples = [
        get_computer_navigation_example().strip(),
        get_computer_multistep_example().strip(),
        get_computer_screenshot_driven_example().strip(),
        get_computer_session_execution_example().strip(),
        get_computer_session_reattachment_example().strip(),
        get_computer_stateful_workflow_example().strip(),
        get_computer_interactive_workflow_example().strip(),
    ]
    return "\n\n".join(examples)


def get_primitives_examples(*, managers: set[str] | None = None) -> str:
    """Get state manager examples, optionally filtered by manager.

    DEPRECATED: Use ToolSurfaceRegistry.prompt_examples(scope) instead.

    Args:
        managers: If provided, only include examples for these managers.
                  If None, include all examples.

    Returns:
        Formatted string with relevant examples.
    """
    from unity.function_manager.primitives import (
        PrimitiveScope,
        VALID_MANAGER_ALIASES,
        get_registry,
    )

    if managers is None:
        scope = PrimitiveScope.all_managers()
    else:
        # Filter to valid aliases only
        valid = managers & VALID_MANAGER_ALIASES
        if not valid:
            return ""
        scope = PrimitiveScope(scoped_managers=frozenset(valid))

    return get_registry().prompt_examples(scope)


def get_mixed_examples() -> str:
    """Get all mixed-mode examples."""

    examples = [
        get_mixed_browse_persist_example().strip(),
        get_mixed_concurrent_example().strip(),
        get_mixed_interjection_routing_example().strip(),
    ]
    return "\n\n".join(examples)


def get_examples_for_environments(
    has_computer: bool,
    has_primitives: bool,
    include_core: bool = True,
    *,
    managers: set[str] | None = None,
) -> str:
    """Get examples appropriate for the given environment combination.

    Args:
        has_computer: Whether primitives.computer environment is active
        has_primitives: Whether primitives environment is active
        include_core: Whether to include core patterns (default: True)
        managers: If provided, only include examples for these managers

    Returns:
        Formatted string with relevant examples
    """

    sections: list[str] = []

    if include_core:
        sections.append(
            "### Core Patterns (Environment-Agnostic)\n" + get_core_pattern_examples(),
        )

    if has_computer:
        sections.append("### Computer Examples\n" + get_computer_examples())

    if has_primitives:
        sections.append(
            "### State Manager Examples\n" + get_primitives_examples(managers=managers),
        )

    if has_computer and has_primitives:
        sections.append("### Mixed-Mode Examples\n" + get_mixed_examples())

    return "\n\n".join(sections)
